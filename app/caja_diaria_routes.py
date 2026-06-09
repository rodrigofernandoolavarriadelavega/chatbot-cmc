"""
Router /alma/api/caja-diaria — Módulo Caja Diaria de Alma.

Réplica del libro de caja manual de recepción (Excel "CAJA DIARIA"):
una fila por día con el efectivo recaudado, comentarios (retiros/ajustes:
"-8550 agua", devoluciones, "sencillo"), y el depósito al banco cuando se hace
(fecha + valor, que suele consolidar varios días).

Columnas (mismo orden que el Excel):
    FECHA · MONTO EFECTIVO · COMENTARIO · FECHA DEPÓSITO · VALOR DEPÓSITO ·
    DÍAS QUE CORRESPONDEN

Conexión con el módulo Pagos: el MONTO EFECTIVO del día = la suma de los pagos
en efectivo de `pagos_cmc` de esa fecha. El GET devuelve ese valor sugerido por
día (`efectivo_pagos`) para que recepción lo traiga de un clic en vez de tipear.

Reconciliación: saldo sin depositar = Σ monto_efectivo − Σ valor_deposito
(el efectivo que aún no se lleva al banco).

Auth y DB: mismo patrón que pagos_routes/abonos_routes (sessions.db,
_require_admin_dep; scope de profesional NO aplica — es caja global del centro).
"""
import io
import logging
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Cookie, HTTPException, Query, Request
from fastapi.responses import StreamingResponse

log = logging.getLogger("caja_diaria_routes")
_CHILE_TZ = ZoneInfo("America/Santiago")

router = APIRouter(prefix="/alma/api/caja-diaria", tags=["caja-diaria"])


# ── Auth (idéntico a pagos_routes/abonos_routes) ─────────────────────────────

def _require_admin_dep(request: Request,
                       token: str | None = Query(None),
                       cmc_session: str | None = Cookie(None)) -> str:
    import hmac as _hmac
    from config import ADMIN_TOKEN
    from admin_routes import _verify_cookie, _is_admin_token

    auth_header = request.headers.get("authorization", "") if request else ""
    if auth_header.lower().startswith("bearer "):
        tk = auth_header.split(None, 1)[1].strip()
        if _is_admin_token(tk):
            return tk
    if cmc_session:
        role = _verify_cookie(cmc_session)
        if role in ("admin", "ortodoncia"):
            return ADMIN_TOKEN
    if token and _is_admin_token(token):
        return token
    raise HTTPException(status_code=401, detail="Token inválido")


# ── DDL ──────────────────────────────────────────────────────────────────────

def ensure_caja_diaria_table() -> None:
    """Crea la tabla caja_diaria si no existe. Idempotente. Una fila por día."""
    from session import _conn
    with _conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS caja_diaria (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha             TEXT NOT NULL UNIQUE,    -- YYYY-MM-DD (día de caja)
                monto_efectivo    INTEGER DEFAULT 0,       -- efectivo recaudado ese día
                comentario        TEXT DEFAULT '',         -- retiros/ajustes/devoluciones
                fecha_deposito    TEXT DEFAULT '',         -- YYYY-MM-DD del depósito al banco
                valor_deposito    INTEGER DEFAULT 0,       -- monto depositado (consolida días)
                dias_corresponden TEXT DEFAULT '',         -- qué días cubre el depósito
                creado_por        TEXT DEFAULT 'recepcion',
                created_at        TEXT DEFAULT (datetime('now')),
                updated_at        TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_caja_diaria_fecha ON caja_diaria(fecha)")


def _vint(v) -> int:
    """Parsea un monto que puede venir como '129.000', '129000', 129000, '' → int."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    # quita separadores de miles (punto/espacio) y símbolos
    s = s.replace("$", "").replace(" ", "").replace(".", "").replace(",", "")
    s = "".join(ch for ch in s if (ch.isdigit() or ch == "-"))
    try:
        return int(s)
    except ValueError:
        return 0


def _vfecha(v) -> str:
    """Normaliza una fecha a YYYY-MM-DD; '' si no parsea."""
    if not v:
        return ""
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


# ── Efectivo sugerido desde el módulo Pagos (pagos_cmc) ──────────────────────

def _efectivo_pagos_por_fecha(d_desde: str, d_hasta: str) -> dict[str, int]:
    """Suma del efectivo registrado en pagos_cmc por fecha en el rango.
    Es el MISMO número que muestra 'efectivo en caja' del módulo Pagos."""
    from session import _conn
    out: dict[str, int] = {}
    try:
        with _conn() as conn:
            for r in conn.execute(
                """SELECT fecha, COALESCE(SUM(copago),0)
                     FROM pagos_cmc
                    WHERE fecha BETWEEN ? AND ? AND metodo_pago = 'efectivo'
                    GROUP BY fecha""",
                (d_desde, d_hasta),
            ):
                out[r[0]] = int(r[1] or 0)
    except Exception as e:
        log.debug("caja_diaria: efectivo_pagos error: %s", e)
    return out


# ── GET: filas del rango + sugerencias + totales ─────────────────────────────

@router.get("")
def get_caja_diaria(
    request: Request,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_caja_diaria_table()

    now_cl = datetime.now(_CHILE_TZ)
    # Por defecto: mes en curso
    d_hasta = _vfecha(fecha_hasta) or now_cl.strftime("%Y-%m-%d")
    d_desde = _vfecha(fecha_desde) or now_cl.strftime("%Y-%m-01")

    from session import _conn
    with _conn() as conn:
        rows = conn.execute(
            """SELECT id, fecha, monto_efectivo, comentario,
                      fecha_deposito, valor_deposito, dias_corresponden
                 FROM caja_diaria
                WHERE fecha BETWEEN ? AND ?
                ORDER BY fecha DESC, id DESC""",
            (d_desde, d_hasta),
        ).fetchall()

    suger = _efectivo_pagos_por_fecha(d_desde, d_hasta)

    filas = []
    total_efectivo = total_deposito = 0
    for r in rows:
        total_efectivo += int(r["monto_efectivo"] or 0)
        total_deposito += int(r["valor_deposito"] or 0)
        filas.append({
            "id":                r["id"],
            "fecha":             r["fecha"],
            "monto_efectivo":    int(r["monto_efectivo"] or 0),
            "comentario":        r["comentario"] or "",
            "fecha_deposito":    r["fecha_deposito"] or "",
            "valor_deposito":    int(r["valor_deposito"] or 0),
            "dias_corresponden": r["dias_corresponden"] or "",
            "efectivo_pagos":    suger.get(r["fecha"], 0),   # sugerido del módulo Pagos
        })

    # Días que tienen efectivo en Pagos pero AÚN no están en el libro de caja
    fechas_libro = {f["fecha"] for f in filas}
    faltantes = [
        {"fecha": f, "efectivo_pagos": v}
        for f, v in sorted(suger.items(), reverse=True)
        if f not in fechas_libro and v > 0
    ]

    return {
        "filas":            filas,
        "total":            len(filas),
        "total_efectivo":   total_efectivo,
        "total_deposito":   total_deposito,
        "saldo_sin_depositar": total_efectivo - total_deposito,
        "faltantes":        faltantes,
        "fecha_desde":      d_desde,
        "fecha_hasta":      d_hasta,
    }


# ── GET /saldo: "lo que hay en la caja en el momento" (vista recepción) ──────
# Efectivo acumulado que aún NO se deposita = Σ efectivo − Σ depósito (hasta hoy).
# El efectivo de HOY se suma desde Pagos aunque recepción aún no lo anote en el libro.

def calcular_saldo_caja() -> dict:
    """Lo que hay en la caja AHORA = Σ efectivo − Σ depósito (hasta hoy), reutilizable
    desde el endpoint, el cruce y el reporte. El libro (Excel) es la autoridad; para
    días recientes en blanco (que el import dejó en 0) se toma el efectivo real de Pagos.
    Incluye días sin depositar desde el último depósito (señal de seguridad)."""
    ensure_caja_diaria_table()
    hoy = datetime.now(_CHILE_TZ).strftime("%Y-%m-%d")
    from session import _conn
    with _conn() as conn:
        libro = {row["fecha"]: int(row["monto_efectivo"] or 0)
                 for row in conn.execute(
                     "SELECT fecha, monto_efectivo FROM caja_diaria WHERE fecha <= ?", (hoy,))}
        libro_deposito = int(conn.execute(
            "SELECT COALESCE(SUM(valor_deposito),0) FROM caja_diaria WHERE fecha <= ?", (hoy,)
        ).fetchone()[0] or 0)
        ult = conn.execute(
            """SELECT fecha, valor_deposito FROM caja_diaria
                WHERE valor_deposito > 0 ORDER BY fecha DESC, id DESC LIMIT 1"""
        ).fetchone()

    from datetime import timedelta as _td
    cutoff = (datetime.now(_CHILE_TZ) - _td(days=30)).strftime("%Y-%m-%d")
    pagos_recientes = _efectivo_pagos_por_fecha(cutoff, hoy)
    extra = sum(v for f, v in pagos_recientes.items() if libro.get(f, 0) == 0)
    en_caja = sum(libro.values()) + extra - libro_deposito

    pagos_hoy    = pagos_recientes.get(hoy, 0)
    efectivo_hoy = libro.get(hoy, 0) or pagos_hoy

    dias_sin_depositar = None
    if ult and ult[0]:
        try:
            d0 = datetime.strptime(ult[0][:10], "%Y-%m-%d").date()
            dias_sin_depositar = (datetime.now(_CHILE_TZ).date() - d0).days
        except ValueError:
            pass

    return {
        "en_caja":          en_caja,
        "efectivo_hoy":     efectivo_hoy,
        "pagos_hoy":        pagos_hoy,
        "hoy_en_libro":     hoy in libro,
        "fecha":            hoy,
        "ultimo_deposito":  {"fecha": ult[0], "valor": int(ult[1] or 0)} if ult else None,
        "dias_sin_depositar": dias_sin_depositar,
    }


@router.get("/saldo")
def get_saldo_actual(
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    return calcular_saldo_caja()


@router.get("/cuadre")
def get_cuadre(
    request: Request,
    fecha: str | None = Query(None, description="YYYY-MM-DD; default ayer"),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    """Cruce caja oficial Medilink × recepción: qué cobró Medilink que recepción NO
    registró (plata posiblemente perdida) + bonif Imed. Read-only."""
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    from cuadre_caja import cruce_dia
    f = _vfecha(fecha) or (datetime.now(_CHILE_TZ) - timedelta(days=1)).strftime("%Y-%m-%d")
    return cruce_dia(f)


# ── POST /deposito: recepción registra un depósito (baja el efectivo en caja) ─
# Se anota en la fila del día de caja (default hoy): SUMA al valor_deposito y fija
# fecha_deposito. No pisa el monto_efectivo. El detalle fino vive en OLACORE.

@router.post("/deposito")
async def registrar_deposito(
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_caja_diaria_table()
    body = await request.json()
    valor = _vint(body.get("valor"))
    if valor <= 0:
        raise HTTPException(400, "El valor del depósito debe ser > 0")
    hoy = datetime.now(_CHILE_TZ).strftime("%Y-%m-%d")
    fecha_caja = _vfecha(body.get("fecha")) or hoy
    fecha_dep  = _vfecha(body.get("fecha_deposito")) or hoy
    dias       = (body.get("dias_corresponden") or "").strip()

    from session import _conn
    with _conn() as conn:
        ex = conn.execute(
            "SELECT id, valor_deposito, dias_corresponden FROM caja_diaria WHERE fecha = ?",
            (fecha_caja,),
        ).fetchone()
        if ex:
            nuevo = int(ex["valor_deposito"] or 0) + valor
            ndias = dias or (ex["dias_corresponden"] or "")
            conn.execute(
                """UPDATE caja_diaria SET valor_deposito = ?, fecha_deposito = ?,
                          dias_corresponden = ?, updated_at = datetime('now') WHERE id = ?""",
                (nuevo, fecha_dep, ndias, ex["id"]),
            )
        else:
            # No hay fila para ese día → la creamos con el efectivo de Pagos como base
            efec = _efectivo_pagos_por_fecha(fecha_caja, fecha_caja).get(fecha_caja, 0)
            conn.execute(
                """INSERT INTO caja_diaria
                      (fecha, monto_efectivo, comentario, fecha_deposito, valor_deposito,
                       dias_corresponden, creado_por, created_at, updated_at)
                   VALUES (?,?,?,?,?,?, 'recepcion', datetime('now'), datetime('now'))""",
                (fecha_caja, efec, "", fecha_dep, valor, dias),
            )
        conn.commit()
    log.info("caja_diaria depósito registrado fecha=%s valor=%d", fecha_caja, valor)
    return {"ok": True, "valor": valor}


# ── POST: crear/actualizar la fila de un día (upsert por fecha) ──────────────

@router.post("")
async def upsert_caja_diaria(
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_caja_diaria_table()
    body = await request.json()

    fecha = _vfecha(body.get("fecha"))
    if not fecha:
        raise HTTPException(400, "fecha debe ser YYYY-MM-DD")

    monto_efectivo    = _vint(body.get("monto_efectivo"))
    comentario        = (body.get("comentario") or "").strip()
    fecha_deposito    = _vfecha(body.get("fecha_deposito"))
    valor_deposito    = _vint(body.get("valor_deposito"))
    dias_corresponden = (body.get("dias_corresponden") or "").strip()

    from session import _conn
    with _conn() as conn:
        cur = conn.execute(
            """INSERT INTO caja_diaria
                  (fecha, monto_efectivo, comentario, fecha_deposito,
                   valor_deposito, dias_corresponden, created_at, updated_at)
               VALUES (?,?,?,?,?,?,datetime('now'),datetime('now'))
               ON CONFLICT(fecha) DO UPDATE SET
                  monto_efectivo    = excluded.monto_efectivo,
                  comentario        = excluded.comentario,
                  fecha_deposito    = excluded.fecha_deposito,
                  valor_deposito    = excluded.valor_deposito,
                  dias_corresponden = excluded.dias_corresponden,
                  updated_at        = datetime('now')""",
            (fecha, monto_efectivo, comentario, fecha_deposito,
             valor_deposito, dias_corresponden),
        )
        conn.commit()
        rid = conn.execute("SELECT id FROM caja_diaria WHERE fecha = ?", (fecha,)).fetchone()[0]
    log.info("caja_diaria upsert fecha=%s efectivo=%d deposito=%d", fecha, monto_efectivo, valor_deposito)
    return {"ok": True, "id": rid, "fecha": fecha}


# ── PATCH: editar campos de una fila por id ──────────────────────────────────

@router.patch("/{row_id}")
async def patch_caja_diaria(
    row_id: int,
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    body = await request.json()

    sets, vals = [], []
    if "monto_efectivo" in body:
        sets.append("monto_efectivo = ?"); vals.append(_vint(body["monto_efectivo"]))
    if "comentario" in body:
        sets.append("comentario = ?"); vals.append((body["comentario"] or "").strip())
    if "fecha_deposito" in body:
        sets.append("fecha_deposito = ?"); vals.append(_vfecha(body["fecha_deposito"]))
    if "valor_deposito" in body:
        sets.append("valor_deposito = ?"); vals.append(_vint(body["valor_deposito"]))
    if "dias_corresponden" in body:
        sets.append("dias_corresponden = ?"); vals.append((body["dias_corresponden"] or "").strip())
    if "fecha" in body:
        nf = _vfecha(body["fecha"])
        if not nf:
            raise HTTPException(400, "fecha inválida")
        sets.append("fecha = ?"); vals.append(nf)
    if not sets:
        raise HTTPException(400, "Nada que actualizar")

    sets.append("updated_at = datetime('now')")
    from session import _conn
    with _conn() as conn:
        try:
            conn.execute(f"UPDATE caja_diaria SET {', '.join(sets)} WHERE id = ?", vals + [row_id])
            conn.commit()
        except Exception as e:
            # choque de UNIQUE(fecha) al renombrar la fecha a un día ya existente
            raise HTTPException(409, f"No se pudo guardar: {e}")
    return {"ok": True, "id": row_id}


# ── DELETE ───────────────────────────────────────────────────────────────────

@router.delete("/{row_id}")
def delete_caja_diaria(
    row_id: int,
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    from session import _conn
    with _conn() as conn:
        conn.execute("DELETE FROM caja_diaria WHERE id = ?", (row_id,))
        conn.commit()
    return {"ok": True}


# ── Export Excel (mismo formato que el libro original) ───────────────────────

@router.get("/export")
def export_caja_diaria(
    request: Request,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_caja_diaria_table()
    import openpyxl
    from openpyxl.styles import Font

    now_cl = datetime.now(_CHILE_TZ)
    d_hasta = _vfecha(fecha_hasta) or now_cl.strftime("%Y-%m-%d")
    d_desde = _vfecha(fecha_desde) or now_cl.strftime("%Y-%m-01")

    from session import _conn
    with _conn() as conn:
        rows = conn.execute(
            """SELECT fecha, monto_efectivo, comentario, fecha_deposito,
                      valor_deposito, dias_corresponden
                 FROM caja_diaria WHERE fecha BETWEEN ? AND ?
                ORDER BY fecha ASC""",
            (d_desde, d_hasta),
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CAJA DIARIA"
    headers = ["FECHA", "MONTO EFECTIVO", "COMENTARIO", "FECHA DEPOSITO",
               "VALOR DEPOSITO", "DIAS QUE CORRESPONDEN"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([
            r["fecha"],
            int(r["monto_efectivo"] or 0) or None,
            r["comentario"] or None,
            r["fecha_deposito"] or None,
            int(r["valor_deposito"] or 0) or None,
            r["dias_corresponden"] or None,
        ])
    widths = [13, 16, 48, 16, 16, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"caja_diaria_{d_desde}_{d_hasta}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
