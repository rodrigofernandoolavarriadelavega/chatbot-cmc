"""
Router /alma/api/pagos — Módulo de Pagos Alma (Fase 1).

Reemplaza el Google Sheet "pacientes diarios".
Cada registro = un pago procesado en recepción (presencial, teléfono o chat).

Principio rector: todo pre-llenado como SUGERENCIA, todo editable.
La recepcionista acepta rápido lo normal y corrige lo que varíe
(con bono los copagos bajan y varían).

Auth: misma que agenda_routes (_require_admin_dep).
DB: sessions.db (tabla pagos_cmc), misma que todo el stack.
"""
import logging
import re
from datetime import datetime, date
from zoneinfo import ZoneInfo

import io
from fastapi import APIRouter, Depends, HTTPException, Request, Query, Cookie
from fastapi.responses import JSONResponse, StreamingResponse

log = logging.getLogger("pagos_routes")
_CHILE_TZ = ZoneInfo("America/Santiago")

router = APIRouter(prefix="/alma/api/pagos", tags=["pagos"])

# ── Arancel Fonasa MLE Nivel 3 (CMC) ─────────────────────────────────────────
# Solo estas áreas aceptan bono Imed (Fonasa MLE).
# El resto es particular.
# Fuente: Arancel MLE 2024/2025, Nivel 3.
_ARANCEL_N3: dict[str, dict] = {
    "Medicina General": {
        "codigo": "01 01 001",
        "total":  15_130,
        "bonif":   7_880,
        "copago":  7_250,
    },
    "Medicina Familiar": {
        "codigo": "01 01 001",
        "total":  15_130,
        "bonif":   7_880,
        "copago":  7_250,
    },
    "Kinesiología": {
        "codigo_sesion": "06 01 105",
        "total_sesion":  11_390,
        "bonif_sesion":   7_830,
        "copago_sesion":  3_560,
        "codigo_eval":   "06 01 101",
        "total_eval":     3_680,
        "bonif_eval":     2_530,
        "copago_eval":    1_150,
        # default = sesión
        "codigo": "06 01 105",
        "total":  11_390,
        "bonif":   7_830,
        "copago":  3_560,
    },
    "Nutrición": {
        "codigo": "26 02 001",
        "total":   9_540,
        "bonif":   4_770,
        "copago":  4_770,
    },
    "Psicología Adulto": {
        "codigo": "09 02 001",
        "total":  20_980,
        "bonif":  14_420,
        "copago":  6_560,
    },
    "Psicología Infantil": {
        "codigo": "09 02 001",
        "total":  20_980,
        "bonif":  14_420,
        "copago":  6_560,
    },
    "Psicología": {
        "codigo": "09 02 001",
        "total":  20_980,
        "bonif":  14_420,
        "copago":  6_560,
    },
}

# Precio PARTICULAR por id_profesional (importado de agenda_routes para no duplicar)
# Se importa lazy para evitar circular imports.
def _precio_particular(id_profesional: int) -> int | None:
    try:
        from agenda_routes import _CAPI_VALUE_BY_PROF
        v = _CAPI_VALUE_BY_PROF.get(id_profesional)
        return int(v) if v is not None else None
    except Exception:
        return None


# ── Auth (mismo patrón que agenda_routes) ────────────────────────────────────

def _require_admin_dep(request: Request,
                       token: str | None = Query(None),
                       cmc_session: str | None = Cookie(None)) -> str:
    import hmac as _hmac
    from config import ADMIN_TOKEN
    from admin_routes import _verify_cookie

    auth_header = request.headers.get("authorization", "")
    if auth_header.lower().startswith("bearer "):
        tk = auth_header.split(None, 1)[1].strip()
        if _hmac.compare_digest(tk, ADMIN_TOKEN):
            return tk
    if cmc_session:
        role = _verify_cookie(cmc_session)
        if role in ("admin", "ortodoncia"):
            return ADMIN_TOKEN
    if token and _hmac.compare_digest(token, ADMIN_TOKEN):
        return token
    raise HTTPException(status_code=401, detail="Token inválido")


def _resolve_pagos(request: Request,
                   token: str | None = None,
                   cmc_session: str | None = None) -> tuple[str, int | None]:
    """Auth de Pagos con scope. (token_efectivo, scope_prof).
    scope None = admin/recepción (lectura + escritura completa).
    scope int  = perfil de profesional → SOLO LECTURA, filtrado a su id_profesional.
    La escritura NO usa este resolver: sigue gateada por _require_admin_dep, que da
    401 a cualquier token de profesional. Read-only enforced en el servidor."""
    from alma_scope import resolve
    return resolve(request, token, cmc_session, "pagos")


# ── DDL helper — llamado en lifespan de main.py (o lazy en primera query) ────

def ensure_pagos_table() -> None:
    """Crea la tabla pagos_cmc si no existe. Idempotente."""
    from session import _conn
    with _conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pagos_cmc (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha               TEXT NOT NULL,
                hora                TEXT NOT NULL,
                paciente_nombre     TEXT NOT NULL DEFAULT '',
                rut                 TEXT DEFAULT '',
                id_profesional      INTEGER,
                profesional         TEXT DEFAULT '',
                area                TEXT DEFAULT '',
                prevision           TEXT DEFAULT 'particular',
                copago              INTEGER DEFAULT 0,
                bonificacion        INTEGER DEFAULT 0,
                metodo_pago         TEXT DEFAULT 'efectivo',
                folio               TEXT DEFAULT '',
                codigo_transferencia TEXT DEFAULT '',
                procedimiento       TEXT DEFAULT '',
                origen              TEXT DEFAULT 'presencial',
                id_cita             TEXT DEFAULT '',
                creado_por          TEXT DEFAULT 'recepcion',
                created_at          TEXT DEFAULT (datetime('now')),
                updated_at          TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_pagos_fecha ON pagos_cmc(fecha)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_pagos_rut ON pagos_cmc(rut)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_pagos_prof ON pagos_cmc(id_profesional)"
        )
        # Columnas añadidas post-creación (idempotente)
        for col_ddl in [
            "ALTER TABLE pagos_cmc ADD COLUMN tipo_bono TEXT DEFAULT ''",
            "ALTER TABLE pagos_cmc ADD COLUMN bloqueado INTEGER NOT NULL DEFAULT 0",
            # Canal + Fuente (2026-06-01): modelo de dos niveles para atribución real
            # canal: presencial | telefono | bot | chat_humano
            # fuente: web | meta | instagram | whatsapp | '' (vacío para presencial/telefono)
            "ALTER TABLE pagos_cmc ADD COLUMN canal TEXT DEFAULT 'presencial'",
            "ALTER TABLE pagos_cmc ADD COLUMN fuente TEXT DEFAULT ''",
            # Confianza del cruce con el bot (2026-06-01):
            #   'exacto'   → match por RUT o teléfono (identidad segura)
            #   'probable' → match solo por nombre completo (puede haber homónimo)
            #   ''         → presencial / sin cruce
            "ALTER TABLE pagos_cmc ADD COLUMN match_confianza TEXT DEFAULT ''",
            # Monto REAL de la atención traído de Medilink por prellenar (cuadre al peso). 0 = sin dato.
            "ALTER TABLE pagos_cmc ADD COLUMN monto_medilink INTEGER DEFAULT 0",
        ]:
            try:
                conn.execute(col_ddl)
            except Exception:
                pass  # columna ya existe
        # Origen de primera mano de cada cita (lo que la recepción marca al agendar
        # desde la Agenda de Alma). Dato directo > inferencia: el cruce de Pagos lo
        # lee PRIMERO. NO alimenta recordatorios (no es citas_bot).
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cita_origen (
                id_cita     TEXT PRIMARY KEY,
                canal       TEXT DEFAULT 'presencial',
                fuente      TEXT DEFAULT '',
                created_at  TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.commit()


def registrar_origen_cita(id_cita: str, canal: str, fuente: str = "") -> None:
    """Persiste el origen de primera mano de una cita agendada desde el panel.
    Idempotente (REPLACE). Llamado por agenda_routes al crear la cita."""
    if not id_cita:
        return
    try:
        ensure_pagos_table()
        from session import _conn
        with _conn() as conn:
            conn.execute(
                """INSERT OR REPLACE INTO cita_origen (id_cita, canal, fuente, created_at)
                   VALUES (?, ?, ?, datetime('now'))""",
                (str(id_cita).strip(), canal or "presencial", fuente or "")
            )
            conn.commit()
    except Exception as e:
        log.warning("registrar_origen_cita id_cita=%s error=%s", id_cita, e)


def _buscar_origen_cita(id_cita: str) -> dict | None:
    """Origen de primera mano de una cita (tabla cita_origen). None si no hay."""
    if not id_cita:
        return None
    try:
        from session import _conn
        with _conn() as conn:
            row = conn.execute(
                "SELECT canal, fuente FROM cita_origen WHERE id_cita = ? LIMIT 1",
                (str(id_cita).strip(),)
            ).fetchone()
            if row:
                return {"canal": row["canal"] or "presencial", "fuente": row["fuente"] or ""}
    except Exception as e:
        log.debug("_buscar_origen_cita id_cita=%s error=%s", id_cita, e)
    return None


# ── Lógica de sugerencia ──────────────────────────────────────────────────────

def _sugerir_copago(area: str, prevision: str, id_profesional: int | None) -> dict:
    """
    Retorna copago_sugerido y bonif_sugerida según área y previsión.
    Siempre como sugerencia — el frontend los muestra editables.
    """
    if prevision == "fonasa" and area in _ARANCEL_N3:
        arancel = _ARANCEL_N3[area]
        return {
            "copago_sugerido":  arancel["copago"],
            "bonif_sugerida":   arancel["bonif"],
            "total_arancel":    arancel["total"],
            "codigo_fonasa":    arancel.get("codigo", ""),
            "fuente_arancel":   "fonasa_n3",
        }
    # Particular: copago = precio del profesional, sin bonificación
    precio = _precio_particular(id_profesional) if id_profesional else None
    return {
        "copago_sugerido":  precio or 0,
        "bonif_sugerida":   0,
        "total_arancel":    precio or 0,
        "codigo_fonasa":    "",
        "fuente_arancel":   "particular",
    }


# ── Derivación canal + fuente desde sessions.db ──────────────────────────────

def _derivar_canal_fuente(telefono_medilink: str) -> tuple[str, str]:
    """
    Dado el teléfono del paciente tal como viene de Medilink, cruza contra
    sessions.db LOCAL (cero requests Medilink extra) y devuelve:
        (canal, fuente)

    Modelo de dos niveles:
      canal: 'presencial' | 'telefono' | 'bot' | 'chat_humano'
      fuente: 'web' | 'meta' | 'instagram' | 'whatsapp' | ''

    Lógica:
    - Sin conversación en sessions.db → presencial / ''
    - Con conversación:
        * canal: HUMAN_TAKEOVER activo en la sesión → 'chat_humano'; sino 'bot'
          (también se detecta si hay evento 'derivado_humano' en conversation_events)
        * fuente: meta_referral presente → 'meta'
                  tag 'referral_source:web_*' → 'web'
                  phone con prefijo 'ig_' → 'instagram'
                  phone con prefijo 'fb_' → 'meta'
                  WhatsApp sin referral → 'whatsapp'

    Si el teléfono es vacío o None, retorna ('presencial', '').
    """
    if not telefono_medilink:
        return ("presencial", "")

    try:
        from session import _conn, normalize_wa_id
        # Normalizar al formato canónico del bot (569XXXXXXXX o ig_*/fb_*)
        phone = normalize_wa_id(str(telefono_medilink).strip())

        with _conn() as conn:
            # 1. Verificar si existe sesión (cualquier estado, incluso IDLE)
            sess_row = conn.execute(
                "SELECT state FROM sessions WHERE phone = ? LIMIT 1",
                (phone,)
            ).fetchone()

        if not sess_row:
            return ("presencial", "")

        state = sess_row["state"] or ""

        # 2. Determinar canal
        # HUMAN_TAKEOVER en estado actual → chat_humano
        # También verifica si hubo un evento 'derivado_humano' reciente (30 días)
        canal = "bot"
        if state == "HUMAN_TAKEOVER":
            canal = "chat_humano"
        else:
            with _conn() as conn:
                ev_row = conn.execute(
                    """SELECT 1 FROM conversation_events
                       WHERE phone = ? AND event = 'derivado_humano'
                       ORDER BY ts DESC LIMIT 1""",
                    (phone,)
                ).fetchone()
            if ev_row:
                canal = "chat_humano"

        # 3. Determinar fuente
        # 3a. Prefijo del phone (IG/FB)
        if phone.startswith("ig_"):
            return (canal, "instagram")
        if phone.startswith("fb_"):
            return (canal, "meta")

        # 3b. meta_referral en la tabla (TTL 168h pero para pagos usamos todo el historial)
        with _conn() as conn:
            mr_row = conn.execute(
                "SELECT 1 FROM meta_referrals WHERE phone = ? LIMIT 1",
                (phone,)
            ).fetchone()
        if mr_row:
            return (canal, "meta")

        # 3c. Tags referral_source:web_*
        with _conn() as conn:
            tag_row = conn.execute(
                "SELECT 1 FROM contact_tags WHERE phone = ? AND tag LIKE 'referral_source:web%' LIMIT 1",
                (phone,)
            ).fetchone()
        if tag_row:
            return (canal, "web")

        # 3d. WhatsApp estándar sin referral
        return (canal, "whatsapp")

    except Exception as e:
        log.debug("_derivar_canal_fuente phone=%s error=%s", telefono_medilink, e)
        return ("presencial", "")


def _buscar_telefono_por_id_cita(id_cita: str) -> str:
    """
    Dado un id_cita, busca el teléfono del paciente en citas_bot (sessions.db).
    citas_bot indexa phone↔id_cita cada vez que el bot confirma un agendamiento.
    Es la fuente más directa: no depende de que Medilink devuelva RUT ni celular
    (GET /citas solo trae id_paciente + nombre_paciente, sin teléfono).
    Retorna el primer teléfono encontrado, o '' si no hay.
    """
    if not id_cita:
        return ""
    try:
        from session import _conn
        with _conn() as conn:
            row = conn.execute(
                "SELECT phone FROM citas_bot WHERE id_cita = ? LIMIT 1",
                (str(id_cita).strip(),)
            ).fetchone()
            if row and row["phone"]:
                return row["phone"]
    except Exception as e:
        log.debug("_buscar_telefono_por_id_cita id_cita=%s error=%s", id_cita, e)
    return ""


def _rut_solo_digitos(rut: str) -> str:
    """Cuerpo+DV de un RUT, sin puntos ni guión, en mayúscula.
    Tolera los formatos mezclados de Medilink ('18.543.529-6') y el canónico
    del bot ('18543529-6') → ambos colapsan a '185435296' para comparar.
    """
    if not rut:
        return ""
    try:
        from medilink import clean_rut
        canon = clean_rut(rut) or rut
    except Exception:
        canon = rut
    return re.sub(r"[^0-9K]", "", canon.upper())


def _normalizar_nombre(nombre: str) -> str:
    """Nombre canónico para comparar: minúscula, sin tildes, espacios colapsados,
    tokens ordenados (tolera 'Silva Mora Luis' vs 'Luis Silva Mora')."""
    if not nombre:
        return ""
    try:
        import unicodedata as _ud
        n = _ud.normalize("NFKD", nombre)
        n = "".join(c for c in n if not _ud.combining(c))
    except Exception:
        n = nombre
    n = re.sub(r"[^a-zA-Z\s]", " ", n).lower()
    tokens = [t for t in n.split() if len(t) > 1]
    return " ".join(sorted(tokens))


def _buscar_telefono_por_rut(rut: str) -> str:
    """Teléfono del paciente por RUT en contact_profiles, normalizando ambos lados
    (sin puntos/guión) para no fallar por el formato de Medilink. '' si no hay."""
    rut_norm = _rut_solo_digitos(rut)
    if not rut_norm:
        return ""
    try:
        from session import _conn
        with _conn() as conn:
            row = conn.execute(
                """SELECT phone FROM contact_profiles
                   WHERE REPLACE(REPLACE(UPPER(COALESCE(rut,'')),'.',''),'-','') = ?
                     AND COALESCE(phone,'') != ''
                   LIMIT 1""",
                (rut_norm,)
            ).fetchone()
            if row and row["phone"]:
                return row["phone"]
    except Exception as e:
        log.debug("_buscar_telefono_por_rut rut=%s error=%s", rut, e)
    return ""


def _buscar_profile_por_nombre(nombre: str) -> tuple[str, str]:
    """Cruce por NOMBRE completo (fallback 'probable'): retorna (telefono, rut)
    del perfil del bot cuyo nombre normalizado calce exacto. Permite además
    rellenar el RUT que Medilink no capturó. ('','') si no hay match único."""
    nombre_norm = _normalizar_nombre(nombre)
    if not nombre_norm or len(nombre_norm) < 6:  # evita matches por nombres muy cortos
        return ("", "")
    try:
        from session import _conn
        with _conn() as conn:
            rows = conn.execute(
                """SELECT phone, rut, nombre FROM contact_profiles
                   WHERE COALESCE(phone,'') != '' AND COALESCE(nombre,'') != ''"""
            ).fetchall()
        matches = [
            r for r in rows if _normalizar_nombre(r["nombre"]) == nombre_norm
        ]
        # Solo si hay UN único perfil con ese nombre (evita ambigüedad de homónimos)
        if len(matches) == 1:
            return (matches[0]["phone"] or "", matches[0]["rut"] or "")
    except Exception as e:
        log.debug("_buscar_profile_por_nombre nombre=%s error=%s", nombre, e)
    return ("", "")


def _resolver_atribucion(
    id_cita: str, rut: str, nombre: str, telefono_cita: str = ""
) -> dict:
    """
    Cascada de identificación cita→paciente del bot, con nivel de confianza:
      0. origen de primera mano (cita_origen, marcado al agendar en el panel) → exacto
      1. teléfono explícito en la cita      → exacto
      2. id_cita en citas_bot (bot agendó)  → exacto
      3. RUT en contact_profiles            → exacto
      4. nombre completo único              → probable (+ rellena RUT si falta)
    Luego deriva canal/fuente con _derivar_canal_fuente sobre el teléfono hallado.
    Devuelve: {canal, fuente, confianza, telefono, rut} (rut posiblemente rellenado).
    """
    rut_out = (rut or "").strip()

    # 0. Origen de primera mano: lo que la recepción marcó al agendar en el panel.
    #    Es dato directo, no inferencia → gana a todo lo demás.
    origen_directo = _buscar_origen_cita(id_cita)
    if origen_directo:
        return {
            "canal": origen_directo["canal"],
            "fuente": origen_directo["fuente"],
            "confianza": "exacto",
            "telefono": "",
            "rut": rut_out,
        }

    telefono = (telefono_cita or "").strip()
    confianza = "exacto" if telefono else ""

    if not telefono and id_cita:
        telefono = _buscar_telefono_por_id_cita(id_cita)
        if telefono:
            confianza = "exacto"
    if not telefono and rut_out:
        telefono = _buscar_telefono_por_rut(rut_out)
        if telefono:
            confianza = "exacto"
    if not telefono and nombre:
        tel_nom, rut_nom = _buscar_profile_por_nombre(nombre)
        if tel_nom:
            telefono = tel_nom
            confianza = "probable"
            if not rut_out and rut_nom:  # rellenar RUT que Medilink no capturó
                rut_out = rut_nom

    canal, fuente = _derivar_canal_fuente(telefono)
    if canal == "presencial":
        confianza = ""  # sin conversación con el bot → no es atribuible
    return {
        "canal": canal,
        "fuente": fuente,
        "confianza": confianza,
        "telefono": telefono,
        "rut": rut_out,
    }


def _canal_a_origen(canal: str) -> str:
    """Mapea canal nuevo → origen legacy para mantener compatibilidad con campos existentes."""
    return {
        "bot":         "chat",
        "chat_humano": "chat",
        "telefono":    "telefono",
        "presencial":  "presencial",
    }.get(canal, "presencial")


# Nombre canónico del profesional SOLO para el módulo de Pagos (no afecta los
# mensajes del bot al paciente). Unifica variantes de Medilink en una sola etiqueta.
_NOMBRE_PROF_PAGOS: dict[int, str] = {
    68: "David Pardo M. Servicio Ecotomografía",
}


def _nombre_profesional_pagos(id_prof, nombre_medilink: str) -> str:
    """Etiqueta única del profesional para Pagos: override por id si existe,
    si no el nombre que vino de Medilink."""
    if id_prof in _NOMBRE_PROF_PAGOS:
        return _NOMBRE_PROF_PAGOS[id_prof]
    return (nombre_medilink or "").strip()


def _convenio_a_prevision(nombre_convenio: str) -> str:
    """Mapea el convenio de la atención Medilink → previsión del módulo de Pagos.
    Convenio con 'fonasa' → fonasa; vacío o cualquier otro → particular.
    (CMC opera Fonasa MLE + particular; no hay isapres en convenio.)"""
    n = (nombre_convenio or "").strip().lower()
    if not n:
        return "particular"
    if "fonasa" in n or "mle" in n or "bono" in n:
        return "fonasa"
    return "particular"


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/arancel")
async def get_arancel(
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
    request: Request = None,
):
    """
    Retorna la tabla de arancel Fonasa N3 y la lista de areas que aceptan bono.
    Util para el frontend (pre-llenar selectores).
    """
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    return {
        "nivel": 3,
        "areas_fonasa": list(_ARANCEL_N3.keys()),
        "aranceles": _ARANCEL_N3,
    }


async def _buscar_cita_paciente_fecha(rut: str, fecha: str) -> dict | None:
    """
    Busca en Medilink la(s) cita(s) del paciente (por RUT) para una fecha exacta.
    Retorna la primera cita encontrada como dict con id, id_profesional, profesional, area,
    prestacion, o None si no hay citas ese día.

    Consultas HTTP: 1 (citas) + 1 opcional (atenciones/detalles si hay id_atencion) — 429-safe.

    La prestación viene de /atenciones/{id_atencion}/detalles → nombre_prestacion.
    Este endpoint solo tiene datos cuando la atención ya fue cerrada en Medilink;
    si aún no está cerrada, prestacion queda vacía (editable en el formulario).
    """
    try:
        from medilink import (
            _get_shared_client, _q, _safe_json, HEADERS,
            PROFESIONALES, _rut_safe,
        )
        from config import MEDILINK_BASE_URL

        # RUT canónico: sin puntos, con guión, mayúsculas
        rut_clean = "".join(c for c in rut.upper() if c.isalnum())
        if len(rut_clean) > 1:
            rut_fmt = rut_clean[:-1] + "-" + rut_clean[-1]
        else:
            rut_fmt = rut_clean

        params = {
            "rut":              {"eq": rut_fmt},
            "fecha":            {"eq": fecha},
            "estado_anulacion": {"eq": 0},
        }
        client = _get_shared_client()
        r = await client.get(
            f"{MEDILINK_BASE_URL}/citas",
            params={"q": _q(params)},
            headers=HEADERS,
            timeout=10,
        )
        if r.status_code != 200:
            log.warning(
                "_buscar_cita_paciente_fecha rut=%s fecha=%s → HTTP %d",
                _rut_safe(rut), fecha, r.status_code,
            )
            return None
        data = _safe_json(r).get("data", [])
        if not data:
            return None
        # Ordenar por hora_inicio y tomar la primera
        data.sort(key=lambda c: c.get("hora_inicio", ""))
        cita = data[0]
        id_prof = cita.get("id_profesional")
        prof_info = PROFESIONALES.get(id_prof, {}) if id_prof else {}

        # ── Prestación desde /atenciones/{id_atencion}/detalles ──────────────
        # La cita lleva id_atencion cuando la atención fue creada en Medilink.
        # El detalle tiene nombre_prestacion cuando la atención está cerrada.
        # Si no está cerrada aún, devolvemos prestacion vacía (editable).
        prestacion = ""
        id_aten = cita.get("id_atencion")
        if id_aten:
            try:
                rd = await client.get(
                    f"{MEDILINK_BASE_URL}/atenciones/{id_aten}/detalles",
                    headers=HEADERS,
                    timeout=8,
                )
                if rd.status_code == 200:
                    detalles = _safe_json(rd).get("data", [])
                    # Recopilar nombres únicos de prestaciones realizadas
                    nombres = []
                    for det in detalles:
                        n = (det.get("nombre_prestacion") or "").strip()
                        if n and n not in nombres:
                            nombres.append(n)
                    prestacion = " / ".join(nombres)
                    if prestacion:
                        log.info(
                            "_buscar_cita_paciente_fecha: prestacion desde detalles "
                            "id_aten=%s → %r", id_aten, prestacion
                        )
            except Exception as e_det:
                log.debug("_buscar_cita_paciente_fecha detalles error id_aten=%s: %s", id_aten, e_det)

        return {
            "id_cita":        str(cita["id"]),
            "id_profesional": id_prof,
            "profesional":    cita.get("nombre_profesional", "") or prof_info.get("nombre", ""),
            "area":           prof_info.get("especialidad", ""),
            "prestacion":     prestacion,
        }
    except Exception as e:
        log.warning("_buscar_cita_paciente_fecha error: %s", e)
        return None


@router.get("/sugerencia")
async def get_sugerencia(
    rut: str | None = Query(None),
    id_cita: str | None = Query(None),
    fecha: str | None = Query(None, description="YYYY-MM-DD; por defecto hoy"),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
    request: Request = None,
):
    """
    Devuelve sugerencias pre-llenadas para el formulario de pago.
    Busca en Medilink con rut o id_cita para obtener paciente y cita.
    Si no viene id_cita pero sí rut+fecha, consulta Medilink para auto-rellenar
    id_cita, profesional y área a partir de la cita del paciente ese día.
    Todo es sugerencia — el frontend lo muestra editable.
    """
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    from medilink import PROFESIONALES
    now_cl = datetime.now(_CHILE_TZ)

    # Fecha del registro (normalizada a YYYY-MM-DD)
    fecha_registro = now_cl.strftime("%Y-%m-%d")
    if fecha:
        try:
            datetime.strptime(fecha, "%Y-%m-%d")
            fecha_registro = fecha
        except ValueError:
            pass  # si viene malformada, usar hoy

    sugerencia: dict = {
        "fecha":           fecha_registro,
        "hora":            now_cl.strftime("%H:%M"),
        "paciente_nombre": "",
        "rut":             rut or "",
        "id_profesional":  None,
        "profesional":     "",
        "area":            "",
        "prestacion":      "",
        "prevision":       "particular",
        "copago_sugerido": 0,
        "bonif_sugerida":  0,
        "total_arancel":   0,
        "codigo_fonasa":   "",
        "fuente_arancel":  "sin_datos",   # renombrado para no colisionar con fuente atribución
        "id_cita":         id_cita or "",
        "origen":          "presencial",
        "canal":           "presencial",
        "fuente":          "",
        "metodo_pago":     "efectivo",
    }

    # Si viene id_cita: buscar datos de la cita en cache local
    if id_cita:
        try:
            from session import _conn
            with _conn() as conn:
                row = conn.execute(
                    """SELECT profesional, especialidad, id_cita,
                              phone, paciente_nombre
                       FROM citas_bot
                       WHERE id_cita = ? LIMIT 1""",
                    (id_cita,)
                ).fetchone()
            if row:
                # Buscar id_profesional por nombre en PROFESIONALES
                prof_nombre = row["profesional"] or ""
                id_prof_match = None
                for pid, pinfo in PROFESIONALES.items():
                    if pinfo["nombre"] == prof_nombre:
                        id_prof_match = pid
                        break
                sugerencia.update({
                    "id_profesional":  id_prof_match,
                    "profesional":     prof_nombre,
                    "area":            row["especialidad"] or "",
                    "id_cita":         id_cita,
                    "paciente_nombre": row["paciente_nombre"] or "",
                })
        except Exception as e:
            log.warning("get_sugerencia: error buscando cita cache: %s", e)

    # Si viene rut: buscar paciente en Medilink
    if rut:
        try:
            from medilink import buscar_paciente
            paciente = await buscar_paciente(rut.strip())
            if paciente:
                nombre = (
                    f"{paciente.get('nombre', '')} {paciente.get('apellidos', '')}".strip()
                )
                sugerencia["paciente_nombre"] = nombre
                sugerencia["rut"] = paciente.get("rut", rut)
                # Previsión desde perfil local si existe
                try:
                    from session import _conn
                    with _conn() as conn:
                        prof_row = conn.execute(
                            "SELECT prevision FROM contact_profiles WHERE rut = ? LIMIT 1",
                            (rut.strip(),)
                        ).fetchone()
                    if prof_row and prof_row["prevision"]:
                        sugerencia["prevision"] = prof_row["prevision"]
                except Exception:
                    pass
        except Exception as e:
            log.warning("get_sugerencia: error buscando paciente Medilink: %s", e)

        # Si no hay id_cita todavía, buscar cita del paciente en esa fecha
        if not sugerencia["id_cita"] and rut:
            cita_dia = await _buscar_cita_paciente_fecha(rut.strip(), fecha_registro)
            if cita_dia:
                sugerencia.update({
                    "id_cita":        cita_dia["id_cita"],
                    "id_profesional": cita_dia["id_profesional"],
                    "profesional":    cita_dia["profesional"],
                    "area":           cita_dia["area"],
                    "prestacion":     cita_dia.get("prestacion", ""),
                })
                log.info(
                    "get_sugerencia: cita auto-rellenada id=%s prof=%s fecha=%s",
                    cita_dia["id_cita"], cita_dia["profesional"], fecha_registro,
                )

    # Calcular copago/bonif sugeridos con los datos que tenemos
    area = sugerencia["area"]
    prevision = sugerencia["prevision"]
    id_prof = sugerencia["id_profesional"]
    if area or id_prof:
        montos = _sugerir_copago(area, prevision, id_prof)
        sugerencia.update(montos)

    # Info del profesional (especialidad si area vacía)
    if sugerencia["id_profesional"] and not sugerencia["area"]:
        pinfo = PROFESIONALES.get(sugerencia["id_profesional"], {})
        sugerencia["area"] = pinfo.get("especialidad", "")

    # Derivar canal + fuente desde sessions.db (cero requests Medilink extra)
    # Buscamos el teléfono del paciente en citas_bot (tiene phone ↔ rut) o contact_profiles
    telefono_paciente = _buscar_telefono_por_rut(sugerencia.get("rut", ""))
    canal_sugerido, fuente_sugerida = _derivar_canal_fuente(telefono_paciente)
    sugerencia["canal"]  = canal_sugerido
    sugerencia["fuente"] = fuente_sugerida
    # Mapear canal → origen legacy para compatibilidad
    sugerencia["origen"] = _canal_a_origen(canal_sugerido)

    return sugerencia


@router.post("")
async def post_pago(
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    """
    Crea un registro de pago. Acepta cualquier valor — el formulario ya
    permitió editar las sugerencias.

    Body JSON: ver campos tabla pagos_cmc.
    """
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Body JSON inválido")

    # Validaciones mínimas
    paciente_nombre = (body.get("paciente_nombre") or "").strip()
    if not paciente_nombre:
        raise HTTPException(400, "paciente_nombre es requerido")

    now_cl = datetime.now(_CHILE_TZ)
    fecha = (body.get("fecha") or now_cl.strftime("%Y-%m-%d")).strip()
    hora  = (body.get("hora")  or now_cl.strftime("%H:%M")).strip()

    # Validar fecha
    try:
        datetime.strptime(fecha, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "fecha debe ser YYYY-MM-DD")

    prevision = (body.get("prevision") or "particular").lower()
    if prevision not in ("fonasa", "particular"):
        prevision = "particular"

    metodo_pago = (body.get("metodo_pago") or "efectivo").lower()
    if metodo_pago not in ("efectivo", "transferencia", "debito", "credito"):
        metodo_pago = "efectivo"

    origen = (body.get("origen") or "presencial").lower()
    if origen not in ("chat", "telefono", "presencial"):
        origen = "presencial"

    canal = (body.get("canal") or "presencial").lower()
    if canal not in ("presencial", "telefono", "bot", "chat_humano"):
        canal = "presencial"

    fuente = (body.get("fuente") or "").lower()
    if fuente not in ("web", "meta", "instagram", "whatsapp", ""):
        fuente = ""

    id_prof_raw = body.get("id_profesional")
    id_profesional = int(id_prof_raw) if id_prof_raw is not None else None

    # Nombre canónico por id (evita fragmentar reportes: "Javiera Burgos Godoy"
    # y "Dra. Javiera Burgos" entran ambos como el nombre maestro del id 55).
    from prof_canon import canonical_name
    profesional_canon = canonical_name(id_profesional, body.get("profesional") or "")

    copago      = int(body.get("copago")      or 0)
    bonificacion = int(body.get("bonificacion") or 0)

    from session import _conn
    with _conn() as conn:
        cur = conn.execute(
            """INSERT INTO pagos_cmc
               (fecha, hora, paciente_nombre, rut, id_profesional, profesional,
                area, prevision, copago, bonificacion, metodo_pago, folio,
                codigo_transferencia, tipo_bono, procedimiento, origen, id_cita,
                creado_por, canal, fuente, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'),datetime('now'))""",
            (
                fecha,
                hora,
                paciente_nombre,
                (body.get("rut") or "").strip(),
                id_profesional,
                profesional_canon,
                (body.get("area") or "").strip(),
                prevision,
                copago,
                bonificacion,
                metodo_pago,
                (body.get("folio") or "").strip(),
                (body.get("codigo_transferencia") or "").strip(),
                (body.get("tipo_bono") or "").strip(),
                (body.get("procedimiento") or "").strip(),
                origen,
                (body.get("id_cita") or "").strip(),
                (body.get("creado_por") or "recepcion").strip(),
                canal,
                fuente,
            )
        )
        new_id = cur.lastrowid
        conn.commit()

    log.info(
        "pagos_routes.post_pago: id=%d paciente=%s fecha=%s %s prof=%s "
        "prevision=%s copago=%d bonif=%d metodo=%s",
        new_id, paciente_nombre, fecha, hora,
        body.get("profesional", ""), prevision, copago, bonificacion, metodo_pago,
    )

    return {"ok": True, "id": new_id}


@router.get("")
async def get_pagos(
    fecha: str | None = Query(None, description="YYYY-MM-DD (omitir = hoy)"),
    fecha_desde: str | None = Query(None, description="YYYY-MM-DD rango inicio"),
    fecha_hasta: str | None = Query(None, description="YYYY-MM-DD rango fin"),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
    request: Request = None,
):
    """
    Lista registros de pago.
    - Sin parámetros: retorna el día de hoy (hora Chile).
    - fecha=YYYY-MM-DD: ese día.
    - fecha_desde + fecha_hasta: rango.
    """
    _tk, _scope = _resolve_pagos(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    now_cl = datetime.now(_CHILE_TZ)

    # Determinar rango
    if fecha_desde and fecha_hasta:
        try:
            datetime.strptime(fecha_desde, "%Y-%m-%d")
            datetime.strptime(fecha_hasta, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "Fechas deben ser YYYY-MM-DD")
        d_desde, d_hasta = fecha_desde, fecha_hasta
    else:
        d = fecha or now_cl.strftime("%Y-%m-%d")
        try:
            datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "fecha debe ser YYYY-MM-DD")
        d_desde = d_hasta = d

    from session import _conn
    with _conn() as conn:
        rows = conn.execute(
            """SELECT id, fecha, hora, paciente_nombre, rut,
                      id_profesional, profesional, area, prevision,
                      copago, bonificacion, metodo_pago, folio,
                      codigo_transferencia, tipo_bono, procedimiento, origen,
                      id_cita, creado_por, created_at, updated_at,
                      COALESCE(bloqueado, 0) as bloqueado,
                      COALESCE(canal, 'presencial') as canal,
                      COALESCE(fuente, '') as fuente,
                      COALESCE(match_confianza, '') as match_confianza,
                      COALESCE(monto_medilink, 0) as monto_medilink
               FROM pagos_cmc
               WHERE fecha BETWEEN ? AND ?""" + (" AND id_profesional = ?" if _scope is not None else "") + """
               ORDER BY fecha DESC, hora DESC""",
            ((d_desde, d_hasta, _scope) if _scope is not None else (d_desde, d_hasta))
        ).fetchall()

    pagos = [dict(r) for r in rows]
    # Etiqueta única de profesional en la vista (unifica variantes residuales en DB)
    for p in pagos:
        ov = _NOMBRE_PROF_PAGOS.get(p.get("id_profesional"))
        if ov:
            p["profesional"] = ov
    total_copago = sum(p["copago"] for p in pagos)

    # Bonif calculada desde arancel N3 por area (ya no se guarda en caja)
    def _bonif_para(area: str, prevision: str) -> int:
        if prevision == "fonasa" and area in _ARANCEL_N3:
            return _ARANCEL_N3[area].get("bonif", 0)
        return 0

    total_bonif = sum(
        _bonif_para(p.get("area", ""), p.get("prevision", ""))
        for p in pagos
    )
    # Total recaudado = solo lo que entro a caja (copago)
    total_recaudado = total_copago

    # ── Arancel total (cuadre con Medilink) ──────────────────────────────────
    # Tres datos por atención: efectivo (caja) + por cobrar Imed (bono+seguro) = arancel total.
    #   (A) arancel_ref  → referencia por tipo de atención (tabla Fonasa N3 / particular=copago).
    #   (B) monto_medilink → valor REAL traído de Medilink por prellenar (cuadre al peso). 0 si no hay.
    # Solo cuenta como pago real si hubo copago>0 o método cargado (los placeholders $0 = cita pendiente
    # NO inflan el arancel).
    def _arancel_ref_para(area: str, prevision: str, copago: int) -> int:
        if prevision == "fonasa" and area in _ARANCEL_N3:
            a = _ARANCEL_N3[area]
            return int(a.get("total") or a.get("total_sesion") or copago)
        return int(copago)  # particular: paga el arancel completo

    total_arancel_ref = 0
    total_arancel_medilink = 0
    for p in pagos:
        cop = p.get("copago", 0) or 0
        pago_real = cop > 0 or (p.get("metodo_pago") or "").strip() != ""
        ar = _arancel_ref_para(p.get("area", ""), p.get("prevision", ""), cop) if pago_real else 0
        p["arancel_ref"] = ar
        mm = int(p.get("monto_medilink") or 0)
        # arancel efectivo a usar: el real de Medilink si existe, si no la referencia
        p["arancel"] = mm if mm > 0 else ar
        total_arancel_ref += ar
        total_arancel_medilink += (mm if mm > 0 else ar)
    # Por cobrar a Imed (bono Fonasa + seguros complementarios) = arancel − efectivo
    total_imed = max(total_arancel_medilink - total_copago, 0)

    # ── Arancel REAL de Medilink (caja oficial bi_pagos_caja, sync nocturno T+1) ──
    # Cuadre AL PESO para días ya sincronizados. n_medilink_caja=0 → aún no sincronizado
    # (ej. el día de hoy): el frontend muestra "estimado, cuadra mañana".
    total_medilink_caja = 0
    n_medilink_caja = 0
    try:
        with _conn() as _cc:
            _rcaja = _cc.execute(
                "SELECT COALESCE(SUM(monto), 0), COUNT(*) FROM bi_pagos_caja WHERE fecha BETWEEN ? AND ?",
                (d_desde, d_hasta),
            ).fetchone()
            total_medilink_caja = int(_rcaja[0] or 0)
            n_medilink_caja = int(_rcaja[1] or 0)
    except Exception as _e_caja:
        log.debug("get_pagos: bi_pagos_caja no disponible: %s", _e_caja)

    # Totales por metodo de pago (para sección Caja)
    totales_por_metodo: dict[str, int] = {
        "efectivo":      0,
        "transferencia": 0,
        "debito":        0,
        "credito":       0,
    }
    for p in pagos:
        metodo = (p.get("metodo_pago") or "efectivo").lower()
        if metodo in totales_por_metodo:
            totales_por_metodo[metodo] += p.get("copago", 0) or 0

    return {
        "pagos":               pagos,
        "total":               len(pagos),
        "total_copago":        total_copago,
        "total_bonif":         total_bonif,       # informativo, calculado desde arancel
        "total_recaudado":     total_recaudado,
        "total_arancel":       total_arancel_medilink,  # referencia/cuadre con Medilink
        "total_arancel_ref":   total_arancel_ref,        # solo referencia por tipo (A)
        "total_imed":          total_imed,                # por cobrar a Imed (bono + seguro)
        "total_medilink_caja": total_medilink_caja,       # EXACTO desde bi_pagos_caja (T+1)
        "n_medilink_caja":     n_medilink_caja,            # 0 = día aún no sincronizado
        "totales_por_metodo":  totales_por_metodo,
        "fecha_desde":         d_desde,
        "fecha_hasta":         d_hasta,
    }


@router.patch("/{pago_id}")
async def patch_pago(
    pago_id: int,
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    """
    Edita un registro existente. Solo los campos enviados en el body se actualizan.
    Útil para corregir errores después de guardar.
    """
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Body JSON inválido")

    if not body:
        raise HTTPException(400, "Body vacío")

    # Campos editables (lista blanca)
    _EDITABLE = {
        "fecha", "hora", "paciente_nombre", "rut", "id_profesional",
        "profesional", "area", "prevision", "copago", "bonificacion",
        "metodo_pago", "folio", "codigo_transferencia", "tipo_bono", "procedimiento",
        "origen", "id_cita", "creado_por", "bloqueado", "canal", "fuente",
    }
    campos = {k: v for k, v in body.items() if k in _EDITABLE}
    if not campos:
        raise HTTPException(400, "Sin campos editables en el body")

    # Validaciones puntuales
    if "prevision" in campos and campos["prevision"] not in ("fonasa", "particular"):
        raise HTTPException(400, "prevision debe ser 'fonasa' o 'particular'")
    if "metodo_pago" in campos and campos["metodo_pago"] not in (
        "efectivo", "transferencia", "debito", "credito"
    ):
        raise HTTPException(400, "metodo_pago inválido")
    if "canal" in campos and campos["canal"] not in (
        "presencial", "telefono", "bot", "chat_humano"
    ):
        raise HTTPException(400, "canal inválido")
    if "fuente" in campos and campos["fuente"] not in (
        "web", "meta", "instagram", "whatsapp", ""
    ):
        raise HTTPException(400, "fuente inválida")
    if "fecha" in campos:
        try:
            datetime.strptime(campos["fecha"], "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "fecha debe ser YYYY-MM-DD")

    set_clause = ", ".join(f"{k} = ?" for k in campos)
    values = list(campos.values()) + [pago_id]

    from session import _conn
    with _conn() as conn:
        # Verificar que el registro existe
        exists = conn.execute(
            "SELECT id FROM pagos_cmc WHERE id = ?", (pago_id,)
        ).fetchone()
        if not exists:
            raise HTTPException(404, f"Pago {pago_id} no encontrado")

        conn.execute(
            f"UPDATE pagos_cmc SET {set_clause}, updated_at = datetime('now') WHERE id = ?",
            values
        )
        conn.commit()

    log.info("pagos_routes.patch_pago: id=%d campos=%s", pago_id, list(campos.keys()))
    return {"ok": True, "id": pago_id, "updated": list(campos.keys())}


@router.get("/export")
async def export_pagos_xlsx(
    fecha: str | None = Query(None, description="YYYY-MM-DD (omitir = hoy)"),
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
    request: Request = None,
):
    """
    Exporta los pagos del período como .xlsx.
    Columnas: HORA, PACIENTE, NOMBRE PROFESIONAL, AREA, PREVISION,
              VALOR, METODO DE PAGO, N° FOLIO, PROCEDIMIENTO
    """
    _tk, _scope = _resolve_pagos(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    now_cl = datetime.now(_CHILE_TZ)

    if fecha_desde and fecha_hasta:
        try:
            datetime.strptime(fecha_desde, "%Y-%m-%d")
            datetime.strptime(fecha_hasta, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "Fechas deben ser YYYY-MM-DD")
        d_desde, d_hasta = fecha_desde, fecha_hasta
    else:
        d = fecha or now_cl.strftime("%Y-%m-%d")
        try:
            datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "fecha debe ser YYYY-MM-DD")
        d_desde = d_hasta = d

    from session import _conn
    with _conn() as conn:
        rows = conn.execute(
            """SELECT hora, paciente_nombre, profesional, area, prevision,
                      copago, bonificacion, metodo_pago, folio,
                      codigo_transferencia, tipo_bono, procedimiento,
                      COALESCE(canal, 'presencial') as canal,
                      COALESCE(fuente, '') as fuente
               FROM pagos_cmc
               WHERE fecha BETWEEN ? AND ?""" + (" AND id_profesional = ?" if _scope is not None else "") + """
               ORDER BY fecha ASC, hora ASC""",
            ((d_desde, d_hasta, _scope) if _scope is not None else (d_desde, d_hasta))
        ).fetchall()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"

    # Fila de título con fecha del período
    periodo = d_desde if d_desde == d_hasta else f"{d_desde} — {d_hasta}"
    ws.merge_cells("A1:L1")
    titulo_cell = ws["A1"]
    titulo_cell.value = f"Pagos CMC — {periodo}"
    titulo_cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    titulo_cell.fill = PatternFill("solid", fgColor="0F3F68")
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Encabezados (fila 2)
    headers = [
        "HORA", "PACIENTE", "NOMBRE PROFESIONAL", "AREA",
        "PREVISION", "PAGO", "METODO DE PAGO", "N° FOLIO",
        "COD. TRANSFERENCIA", "TIPO DE BONO", "PRESTACION",
        "CANAL", "FUENTE",
    ]
    header_fill  = PatternFill("solid", fgColor="1172AB")
    header_font  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    thin_border  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[2].height = 16

    # Datos (desde fila 3)
    alt_fill = PatternFill("solid", fgColor="EBF5FB")
    data_font = Font(name="Calibri", size=10)
    for row_idx, row in enumerate(rows, start=3):
        # PAGO = lo que entro a caja (copago solamente)
        valor = row["copago"] or 0
        prevision_label = "Fonasa MLE" if (row["prevision"] or "").lower() == "fonasa" else "Particular"
        metodo_label    = (row["metodo_pago"] or "efectivo").capitalize()
        values = [
            (row["hora"] or "")[:5],
            row["paciente_nombre"] or "",
            row["profesional"] or "",
            row["area"] or "",
            prevision_label,
            valor,
            metodo_label,
            row["folio"] or "",
            row["codigo_transferencia"] or "",
            row["tipo_bono"] or "",
            row["procedimiento"] or "",
            (row["canal"] or "presencial").capitalize(),
            row["fuente"] or "",
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            # Columna VALOR → formato moneda CLP
            if col_idx == 6 and isinstance(v, int):
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Anchos de columna (aproximados)
    col_widths = [8, 24, 22, 18, 12, 12, 14, 12, 18, 14, 28, 12, 12]
    for col_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = w

    # Congelar fila de encabezados
    ws.freeze_panes = "A3"

    # Serializar a bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    label = d_desde if d_desde == d_hasta else f"{d_desde}_{d_hasta}"
    filename = f"pagos_{label}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Mapa de bonificación Fonasa MLE N3 por id_profesional ───────────────────
# Solo profesionales que operan bajo Fonasa MLE N3 en CMC.
# Especialistas (ORL, Cardio, Gastro, Traumato, Gineco) atienden Fonasa como
# libre elección SIN bono Imed → NO se incluyen aquí.
_BONIF_N3_POR_PROF: dict[int, int] = {
    1:  7_880,   # Olavarría MG
    73: 7_880,   # Abarca MG
    13: 7_880,   # Márquez MG/MedFam
    77: 7_830,   # Armijo Kine
    21: 7_830,   # Etcheverry Kine
    52: 4_770,   # Pinto Nutri
    74: 14_420,  # Montalba Psico
    49: 14_420,  # J.P. Rodríguez Psico
}

# Mapa para normalizar método de pago a clave canónica
_METODO_CANON: dict[str, str] = {
    "efectivo": "efectivo",
    "transferencia": "transferencia",
    "debito": "debito",
    "débito": "debito",
    "credito": "credito",
    "crédito": "credito",
    "tarjeta débito": "debito",
    "tarjeta debito": "debito",
    "tarjeta crédito": "credito",
    "tarjeta credito": "credito",
    "tarjeta": "credito",
}


def _normalizar_metodo(metodo_raw: str) -> str:
    """Colapsa variantes de metodo_pago a {efectivo, transferencia, debito, credito, otro}."""
    m = (metodo_raw or "").strip().lower()
    return _METODO_CANON.get(m, "otro")


@router.get("/comision-mes")
async def get_comision_mes(
    mes: str | None = Query(None, description="YYYY-MM; None = último mes con datos"),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
    request: Request = None,
):
    """
    Resumen financiero de un mes para el Simulador de Comisión por Tarjeta.

    Fuentes:
    - caja_total / n_tx / ticket / meses_disponibles: bi_pagos_caja
    - particular / fonasa_copago / bonif_imed: bi_pagos_caja LEFT JOIN bi_atenciones
    - metodo_mix: pagos_cmc (todos los registros disponibles, base pequeña)

    Robustez: cualquier error en BI → ceros + source_status "no_data", nunca 500.
    """
    _require_admin_dep(request, token=token, cmc_session=cmc_session)

    _CERO = {
        "mes": mes or "",
        "meses_disponibles": [],
        "caja_total": 0,
        "n_tx": 0,
        "ticket_promedio": 0,
        "particular": 0,
        "fonasa_copago": 0,
        "n_sin_atencion": 0,
        "bonif_imed": 0,
        "ingreso_total": 0,
        "metodo_mix": {
            "efectivo": 0.0, "transferencia": 0.0, "debito": 0.0,
            "credito": 0.0, "otro": 0.0,
            "pct_migrable": 0.0, "pct_ya_tarjeta": 0.0, "cred_share_en_tarjeta": 0.0,
            "n_registros": 0, "rango": "",
        },
        "tasas": {"debito": 0.006, "credito": 0.013},
        "iva": 0.19,
        "source_status": "no_data",
    }

    try:
        from session import _conn

        # ── 1. Meses disponibles ──────────────────────────────────────────────
        with _conn() as conn:
            rows_meses = conn.execute(
                """SELECT DISTINCT substr(fecha, 1, 7) AS m
                   FROM bi_pagos_caja
                   WHERE fecha >= '2024-01-01'
                   ORDER BY m DESC"""
            ).fetchall()

        meses_disponibles = [r[0] for r in rows_meses]

        if not meses_disponibles:
            return JSONResponse(_CERO)

        # ── 2. Resolver mes objetivo ──────────────────────────────────────────
        if mes and len(mes) == 7:
            mes_obj = mes
        else:
            mes_obj = meses_disponibles[0]  # el más reciente

        yr, mo = int(mes_obj[:4]), int(mes_obj[5:7])
        inicio = f"{mes_obj}-01"
        fin_yr = yr + (mo // 12)
        fin_mo = (mo % 12) + 1
        fin = f"{fin_yr}-{fin_mo:02d}-01"

        # ── 3. caja_total, n_tx, ticket ──────────────────────────────────────
        with _conn() as conn:
            row_caja = conn.execute(
                """SELECT COALESCE(SUM(monto), 0) AS caja,
                          COUNT(*) AS n_tx
                   FROM bi_pagos_caja
                   WHERE fecha >= ? AND fecha < ?""",
                (inicio, fin),
            ).fetchone()

        caja_total = int(row_caja[0])
        n_tx = int(row_caja[1])
        ticket_promedio = (caja_total // n_tx) if n_tx else 0

        if caja_total == 0:
            resp = dict(_CERO)
            resp["mes"] = mes_obj
            resp["meses_disponibles"] = meses_disponibles
            return JSONResponse(resp)

        # ── 4. Particular / Fonasa copago / n_sin_atencion ───────────────────
        # LEFT JOIN bi_atenciones; clasifica por nombre_convenio.
        # Pagos sin fila en bi_atenciones (n_sin_atencion) → particular.
        with _conn() as conn:
            rows_conv = conn.execute(
                """SELECT
                       p.monto,
                       p.id_profesional,
                       LOWER(COALESCE(a.nombre_convenio, '')) AS convenio
                   FROM bi_pagos_caja p
                   LEFT JOIN bi_atenciones a ON p.atencion_id = a.atencion_id
                   WHERE p.fecha >= ? AND p.fecha < ?""",
                (inicio, fin),
            ).fetchall()

        particular = 0
        fonasa_copago = 0
        bonif_imed = 0
        n_sin_atencion = 0

        for monto, id_prof, convenio in rows_conv:
            monto = int(monto or 0)
            es_fonasa = any(kw in convenio for kw in ("fonasa", "mle", "bono"))
            if not convenio:
                # Sin atención enlazada → particular + conteo
                n_sin_atencion += 1
                particular += monto
            elif es_fonasa:
                fonasa_copago += monto
                # Bono Imed solo para profesionales MLE N3
                bonif = _BONIF_N3_POR_PROF.get(id_prof)
                if bonif:
                    bonif_imed += bonif
            else:
                particular += monto

        ingreso_total = caja_total + bonif_imed

        # ── 5. Metodo mix desde pagos_cmc (TODOS los registros disponibles) ──
        with _conn() as conn:
            rows_pago = conn.execute(
                """SELECT copago, metodo_pago,
                          MIN(fecha) AS f_min, MAX(fecha) AS f_max
                   FROM pagos_cmc
                   WHERE COALESCE(copago, 0) > 0
                   GROUP BY id""",  # cada fila es un registro
            ).fetchall()

        # Reagrupar sobre copago total por método
        mix_totales: dict[str, int] = {
            "efectivo": 0, "transferencia": 0, "debito": 0, "credito": 0, "otro": 0,
        }
        f_min_list: list[str] = []
        f_max_list: list[str] = []
        n_registros = 0

        # Necesitamos fecha min/max: re-query sin GROUP BY id
        with _conn() as conn:
            rows_pago = conn.execute(
                """SELECT copago, metodo_pago, fecha
                   FROM pagos_cmc
                   WHERE COALESCE(copago, 0) > 0"""
            ).fetchall()

        for copago, metodo_raw, fecha_p in rows_pago:
            copago = int(copago or 0)
            canon = _normalizar_metodo(metodo_raw or "")
            mix_totales[canon] = mix_totales.get(canon, 0) + copago
            if fecha_p:
                f_min_list.append(fecha_p)
                f_max_list.append(fecha_p)
            n_registros += 1

        total_mix = sum(mix_totales.values()) or 1  # evitar div/0

        # Proporciones sobre suma de copago (no por conteo)
        efectivo_pct     = mix_totales["efectivo"]     / total_mix
        transferencia_pct = mix_totales["transferencia"] / total_mix
        debito_pct       = mix_totales["debito"]       / total_mix
        credito_pct      = mix_totales["credito"]      / total_mix
        otro_pct         = mix_totales["otro"]         / total_mix

        ya_tarjeta = debito_pct + credito_pct
        migrable   = efectivo_pct + transferencia_pct
        cred_share_en_tarjeta = (credito_pct / ya_tarjeta) if ya_tarjeta > 0 else 0.0

        rango = ""
        if f_min_list and f_max_list:
            rango = f"{min(f_min_list)} a {max(f_max_list)}"

        metodo_mix = {
            "efectivo":             round(efectivo_pct, 4),
            "transferencia":        round(transferencia_pct, 4),
            "debito":               round(debito_pct, 4),
            "credito":              round(credito_pct, 4),
            "otro":                 round(otro_pct, 4),
            "pct_migrable":         round(migrable, 4),
            "pct_ya_tarjeta":       round(ya_tarjeta, 4),
            "cred_share_en_tarjeta": round(cred_share_en_tarjeta, 4),
            "n_registros":          n_registros,
            "rango":                rango,
        }

        return JSONResponse({
            "mes":               mes_obj,
            "meses_disponibles": meses_disponibles,
            "caja_total":        caja_total,
            "n_tx":              n_tx,
            "ticket_promedio":   ticket_promedio,
            "particular":        particular,
            "fonasa_copago":     fonasa_copago,
            "n_sin_atencion":    n_sin_atencion,
            "bonif_imed":        bonif_imed,
            "ingreso_total":     ingreso_total,
            "metodo_mix":        metodo_mix,
            "tasas":             {"debito": 0.006, "credito": 0.013},
            "iva":               0.19,
            "source_status":     "ok",
        })

    except Exception as exc:
        log.warning("get_comision_mes error: %s", exc)
        resp = dict(_CERO)
        resp["mes"] = mes or ""
        resp["meses_disponibles"] = []
        resp["source_status"] = "error"
        return JSONResponse(resp)


@router.post("/cruzar-origen")
async def cruzar_origen(
    fecha: str | None = Query(None, description="YYYY-MM-DD; por defecto hoy"),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
    request: Request = None,
):
    """
    Recalcula el ORIGEN (canal/fuente) de las filas de pagos_cmc cruzando contra
    el bot (sessions.db). NO consulta Medilink → instantáneo, idempotente, 429-safe.

    Para cada fila NO bloqueada de la fecha:
      - resuelve atribución por id_cita → teléfono → RUT (exacto) o nombre (probable)
      - actualiza canal, fuente, origen (legacy) y match_confianza
      - rellena el RUT si estaba vacío y el bot lo conoce

    Respeta filas bloqueadas y NO toca copago/método/folio/prestación.
    Retorna: { revisadas, exacto, probable, presencial, rut_rellenados }
    """
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    now_cl = datetime.now(_CHILE_TZ)
    if fecha:
        try:
            datetime.strptime(fecha, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "fecha debe ser YYYY-MM-DD")
        fecha_iso = fecha
    else:
        fecha_iso = now_cl.strftime("%Y-%m-%d")

    from session import _conn
    with _conn() as conn:
        rows = conn.execute(
            """SELECT id, id_cita, rut, paciente_nombre,
                      COALESCE(bloqueado, 0) AS bloqueado
               FROM pagos_cmc WHERE fecha = ?""",
            (fecha_iso,)
        ).fetchall()

    revisadas = exacto = probable = presencial = rut_rellenados = 0
    for row in rows:
        if row["bloqueado"]:
            continue
        revisadas += 1
        att = _resolver_atribucion(
            id_cita=row["id_cita"] or "",
            rut=row["rut"] or "",
            nombre=row["paciente_nombre"] or "",
        )
        origen_legacy = _canal_a_origen(att["canal"])
        nuevo_rut = att["rut"]
        relleno = bool(nuevo_rut and not (row["rut"] or "").strip())

        if att["confianza"] == "exacto":
            exacto += 1
        elif att["confianza"] == "probable":
            probable += 1
        else:
            presencial += 1
        if relleno:
            rut_rellenados += 1

        try:
            with _conn() as conn:
                conn.execute(
                    """UPDATE pagos_cmc
                       SET canal = ?, fuente = ?, origen = ?, match_confianza = ?,
                           rut = CASE WHEN COALESCE(rut,'') = '' THEN ? ELSE rut END,
                           updated_at = datetime('now')
                       WHERE id = ? AND COALESCE(bloqueado, 0) = 0""",
                    (att["canal"], att["fuente"], origen_legacy, att["confianza"],
                     nuevo_rut, row["id"])
                )
                conn.commit()
        except Exception as e_upd:
            log.warning("cruzar_origen: error UPDATE id=%s: %s", row["id"], e_upd)

    log.info(
        "cruzar_origen: fecha=%s revisadas=%d exacto=%d probable=%d presencial=%d rut_rellenados=%d",
        fecha_iso, revisadas, exacto, probable, presencial, rut_rellenados,
    )
    return {
        "revisadas": revisadas, "exacto": exacto, "probable": probable,
        "presencial": presencial, "rut_rellenados": rut_rellenados,
    }


@router.post("/prellenar")
async def prellenar_pagos(
    fecha: str | None = Query(None, description="YYYY-MM-DD; por defecto hoy"),
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
    request: Request = None,
):
    """
    Actualiza incrementalmente las filas de pagos_cmc desde Medilink.

    Estrategia (idempotente, 429-safe):
    - 1 sola query GET /citas?fecha=... para todas las citas del día.
    - Para citas NUEVAS (no existen en pagos_cmc): INSERT draft.
    - Para citas YA EXISTENTES con procedimiento vacío y fila NO bloqueada:
      busca /atenciones/{id}/detalles y rellena el hueco. Throttled 0.15s.
    - NUNCA sobrescribe filas bloqueadas (bloqueado=1) ni campos ya editados
      (copago, metodo_pago, folio, procedimiento no vacíos).
    - Solo escribe procedimiento en filas cuya prestación sigue vacía.

    Retorna: { creadas: N, actualizadas: N, saltadas: N, errores: N }
    """
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    import asyncio
    from medilink import (
        _get_shared_client, _q, _safe_json, HEADERS,
        PROFESIONALES,
    )
    from config import MEDILINK_BASE_URL

    now_cl = datetime.now(_CHILE_TZ)
    if fecha:
        try:
            datetime.strptime(fecha, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "fecha debe ser YYYY-MM-DD")
        fecha_iso = fecha
    else:
        fecha_iso = now_cl.strftime("%Y-%m-%d")

    # ── 1 request: todas las citas del día ───────────────────────────────────
    client = _get_shared_client()
    try:
        r = await client.get(
            f"{MEDILINK_BASE_URL}/citas",
            params={"q": _q({
                "fecha":            {"eq": fecha_iso},
                "estado_anulacion": {"eq": 0},
            })},
            headers=HEADERS,
            timeout=15,
        )
    except Exception as e:
        log.error("prellenar_pagos: error GET /citas fecha=%s: %s", fecha_iso, e)
        raise HTTPException(502, "Error al contactar Medilink")

    if r.status_code != 200:
        log.warning("prellenar_pagos: GET /citas HTTP %d fecha=%s", r.status_code, fecha_iso)
        raise HTTPException(502, f"Medilink devolvió HTTP {r.status_code}")

    citas = _safe_json(r).get("data", [])
    if not citas:
        return {"creadas": 0, "actualizadas": 0, "saltadas": 0, "errores": 0, "mensaje": "Sin citas para esa fecha en Medilink"}

    citas.sort(key=lambda c: c.get("hora_inicio", ""))

    # ── Cargar filas existentes en pagos_cmc para esa fecha ──────────────────
    # Necesitamos: id_cita, rut, procedimiento actual, bloqueado, canal, fuente y el id de la fila
    from session import _conn
    with _conn() as conn:
        rows_exist = conn.execute(
            """SELECT id, id_cita, rut,
                      COALESCE(procedimiento, '') as procedimiento,
                      COALESCE(bloqueado, 0) as bloqueado,
                      COALESCE(canal, 'presencial') as canal,
                      COALESCE(fuente, '') as fuente,
                      COALESCE(match_confianza, '') as match_confianza,
                      COALESCE(copago, 0) as copago,
                      COALESCE(metodo_pago, '') as metodo_pago,
                      COALESCE(creado_por, '') as creado_por,
                      COALESCE(area, '') as area,
                      COALESCE(prevision, 'particular') as prevision,
                      COALESCE(profesional, '') as profesional,
                      COALESCE(monto_medilink, 0) as monto_medilink,
                      id_profesional
               FROM pagos_cmc WHERE fecha = ?""",
            (fecha_iso,)
        ).fetchall()

    # Mapas para búsqueda rápida
    # id_cita → {db_id, procedimiento, bloqueado, canal, fuente}
    existing_by_id_cita: dict[str, dict] = {}
    # rut → {db_id, procedimiento, bloqueado, canal, fuente}
    existing_by_rut: dict[str, dict] = {}
    for row in rows_exist:
        meta = {
            "db_id":         row["id"],
            "procedimiento": row["procedimiento"],
            "bloqueado":     row["bloqueado"],
            "canal":         row["canal"],
            "fuente":        row["fuente"],
            "match_confianza": row["match_confianza"],
            "rut":           row["rut"] or "",
            "copago":        row["copago"],
            "metodo_pago":   row["metodo_pago"],
            "creado_por":    row["creado_por"],
            "area":          row["area"],
            "prevision":     row["prevision"],
            "profesional":   row["profesional"],
            "monto_medilink": row["monto_medilink"],
            "id_profesional": row["id_profesional"],
        }
        if row["id_cita"]:
            existing_by_id_cita[row["id_cita"]] = meta
        if row["rut"]:
            existing_by_rut[row["rut"]] = meta

    creadas = actualizadas = saltadas = errores = 0

    async def _fetch_prestacion(id_aten) -> str:
        """Obtiene prestaciones desde /atenciones/{id}/detalles. Throttleado."""
        try:
            await asyncio.sleep(0.15)
            rd = await client.get(
                f"{MEDILINK_BASE_URL}/atenciones/{id_aten}/detalles",
                headers=HEADERS,
                timeout=8,
            )
            if rd.status_code == 200:
                detalles = _safe_json(rd).get("data", [])
                nombres = []
                for det in detalles:
                    n = (det.get("nombre_prestacion") or "").strip()
                    if n and n not in nombres:
                        nombres.append(n)
                return " / ".join(nombres)
        except Exception as e_det:
            log.debug("prellenar_pagos: detalles id_aten=%s error: %s", id_aten, e_det)
        return ""

    _aten_meta_cache: dict = {}

    async def _fetch_atencion_meta(id_aten) -> tuple[str, int]:
        """Trae (nombre_convenio, total) desde /atenciones/{id}. El convenio define
        la previsión (Fonasa vs particular); total es el arancel. Cacheado + throttled."""
        if not id_aten:
            return ("", 0)
        if id_aten in _aten_meta_cache:
            return _aten_meta_cache[id_aten]
        res = ("", 0)
        try:
            await asyncio.sleep(0.15)
            ra = await client.get(
                f"{MEDILINK_BASE_URL}/atenciones/{id_aten}",
                headers=HEADERS, timeout=8,
            )
            if ra.status_code == 200:
                at = _safe_json(ra).get("data", {}) or {}
                res = ((at.get("nombre_convenio") or "").strip(), int(at.get("total") or 0))
        except Exception as e_at:
            log.debug("prellenar_pagos: atencion id=%s error: %s", id_aten, e_at)
        _aten_meta_cache[id_aten] = res
        return res

    _ficha_cache: dict = {}

    async def _fetch_ficha(id_paciente) -> tuple[str, str]:
        """Trae (rut, telefono) desde la ficha /pacientes/{id} de Medilink.
        GET /citas no incluye rut/celular, pero la ficha sí. Cacheado + throttled.
        Esta es la fuente autoritativa del RUT y el puente para atribución exacta."""
        if not id_paciente:
            return ("", "")
        if id_paciente in _ficha_cache:
            return _ficha_cache[id_paciente]
        rut_tel = ("", "")
        try:
            await asyncio.sleep(0.18)
            rp = await client.get(
                f"{MEDILINK_BASE_URL}/pacientes/{id_paciente}",
                headers=HEADERS, timeout=8,
            )
            if rp.status_code == 200:
                ficha = _safe_json(rp).get("data", {}) or {}
                rut = (ficha.get("rut") or "").strip()
                tel = (ficha.get("celular") or ficha.get("telefono") or "").strip()
                rut_tel = (rut, tel)
        except Exception as e_fic:
            log.debug("prellenar_pagos: ficha id_paciente=%s error: %s", id_paciente, e_fic)
        _ficha_cache[id_paciente] = rut_tel
        return rut_tel

    def _telefono_para_cita(id_cita_str: str, rut_cita: str) -> str:
        """
        Retorna el teléfono del paciente para derivar canal/fuente.
        Orden de prioridad (de más a menos confiable):
          1. citas_bot.phone por id_cita   ← el puente correcto
                GET /citas NO devuelve celular/rut_paciente; solo id_paciente + nombre_paciente.
                citas_bot tiene id_cita ↔ phone para todas las citas agendadas por el bot.
          2. contact_profiles.phone por rut ← fallback si el bot no agendó la cita
        """
        phone = _buscar_telefono_por_id_cita(id_cita_str)
        if not phone and rut_cita:
            phone = _buscar_telefono_por_rut(rut_cita)
        return phone

    for cita in citas:
        id_cita_str = str(cita.get("id", ""))
        # GET /citas no devuelve rut/celular: solo id_paciente + nombre_paciente.
        # La ficha /pacientes/{id} SÍ los trae → fuente autoritativa del RUT y
        # puente para atribución EXACTA por teléfono.
        rut_cita    = (cita.get("rut_paciente") or "").strip()
        id_aten     = cita.get("id_atencion")
        id_paciente = cita.get("id_paciente")

        # Buscar si ya existe una fila para esta cita (por id_cita; rut viene vacío de /citas)
        existing = existing_by_id_cita.get(id_cita_str) if id_cita_str else None

        # Pedir la ficha (rut+tel) SOLO si hace falta — evita 504 por timeout en
        # re-ejecuciones: si la fila ya tiene RUT y canal resuelto, no se consulta.
        need_ficha = (
            existing is None
            or not (existing.get("rut") or "").strip()
            or (existing.get("canal") or "presencial") == "presencial"
        )
        if need_ficha:
            rut_ficha, tel_ficha = await _fetch_ficha(id_paciente)
        else:
            rut_ficha, tel_ficha = "", ""
        if rut_ficha and not rut_cita:
            rut_cita = rut_ficha
        if existing is None and rut_cita:
            existing = existing_by_rut.get(rut_cita)

        if existing:
            # Fila ya existe y no está bloqueada → re-derivar canal/fuente + rellenar
            # hueco de prestación si corresponde.
            if existing["bloqueado"]:
                saltadas += 1
                continue

            # Re-derivar atribución con el teléfono de la ficha (cruce EXACTO).
            # Si no se pidió ficha (fila ya resuelta), mantener su canal/fuente para
            # NO degradar a presencial en re-ejecuciones.
            if need_ficha:
                att = _resolver_atribucion(
                    id_cita=id_cita_str, rut=rut_cita,
                    nombre=cita.get("nombre_paciente") or "", telefono_cita=tel_ficha,
                )
                canal_nuevo, fuente_nueva = att["canal"], att["fuente"]
                confianza_nueva = att["confianza"]
            else:
                canal_nuevo, fuente_nueva = existing["canal"], existing["fuente"]
                confianza_nueva = existing.get("match_confianza") or ""
            origen_nuevo = _canal_a_origen(canal_nuevo)
            # Rellenar RUT desde la ficha si la fila no lo tenía
            rut_relleno = rut_ficha if (rut_ficha and not (existing.get("rut") or "").strip()) else None

            # Rellenar hueco de prestación si corresponde
            prestacion_update = None
            if not existing["procedimiento"] and id_aten:
                prestacion_update = await _fetch_prestacion(id_aten)

            # Previsión + precio desde la atención SOLO si la fila sigue draft
            # (copago 0, sin método, creada por prellenar) → no pisa ediciones de recepción.
            prev_update = None
            copago_update = None
            monto_update = None
            es_draft = (
                existing.get("copago", 0) == 0
                and not (existing.get("metodo_pago") or "").strip()
                and existing.get("creado_por") == "prellenar"
            )
            _falta_monto = bool(id_aten) and not existing.get("monto_medilink")
            if id_aten and (es_draft or _falta_monto):
                convenio, _total = await _fetch_atencion_meta(id_aten)
                if _total and int(_total) > 0 and not existing.get("monto_medilink"):
                    monto_update = int(_total)   # arancel real Medilink (Fase B + backfill filas llenas)
                if es_draft:
                    prev_calc = _convenio_a_prevision(convenio)
                    sug = _sugerir_copago(existing.get("area") or "", prev_calc, existing.get("id_profesional"))
                    copago_calc = int(sug.get("copago_sugerido") or 0)
                    if prev_calc != (existing.get("prevision") or "particular"):
                        prev_update = prev_calc
                    if copago_calc and copago_calc != existing.get("copago", 0):
                        copago_update = copago_calc
                        if prev_update is None:  # si cambia el copago, fijar también la previsión coherente
                            prev_update = prev_calc

            # Normalizar nombre del profesional (etiqueta única en Pagos)
            id_prof_cita = cita.get("id_profesional")
            prof_norm = _nombre_profesional_pagos(id_prof_cita, cita.get("nombre_profesional") or existing.get("profesional") or "")
            prof_update = prof_norm if (prof_norm and prof_norm != (existing.get("profesional") or "")) else None

            # Solo actualizar si hubo cambio real (diff) para que el contador sea honesto
            canal_cambio     = canal_nuevo     != existing["canal"]
            fuente_cambio    = fuente_nueva    != existing["fuente"]
            confianza_cambio = confianza_nueva != (existing.get("match_confianza") or "")
            hay_cambio       = (canal_cambio or fuente_cambio or confianza_cambio or bool(prestacion_update)
                                or bool(rut_relleno) or bool(prev_update) or bool(copago_update) or bool(prof_update)
                                or bool(monto_update))

            if not hay_cambio:
                saltadas += 1
                continue

            # Decidir qué actualizar
            set_parts = ["canal = ?", "fuente = ?", "origen = ?", "match_confianza = ?", "updated_at = datetime('now')"]
            set_vals: list = [canal_nuevo, fuente_nueva, origen_nuevo, confianza_nueva]

            if rut_relleno:
                set_parts.insert(0, "rut = ?")
                set_vals.insert(0, rut_relleno)

            if prof_update is not None:
                set_parts.insert(len(set_parts)-1, "profesional = ?")
                set_vals.append(prof_update)

            if prev_update is not None:
                set_parts.insert(len(set_parts)-1, "prevision = ?")
                set_vals.append(prev_update)
            if copago_update is not None:
                set_parts.insert(len(set_parts)-1, "copago = ?")
                set_vals.append(copago_update)
            if monto_update is not None:
                set_parts.insert(len(set_parts)-1, "monto_medilink = ?")
                set_vals.append(monto_update)

            if prestacion_update:
                set_parts.insert(len(set_parts)-1, "procedimiento = ?")
                set_vals.insert(len(set_vals), prestacion_update)

            try:
                with _conn() as conn:
                    conn.execute(
                        "UPDATE pagos_cmc SET " + ", ".join(set_parts) +
                        " WHERE id = ? AND COALESCE(bloqueado, 0) = 0",
                        set_vals + [existing["db_id"]]
                    )
                    conn.commit()
                actualizadas += 1
                log.info(
                    "prellenar_pagos: actualizado id=%d canal=%s fuente=%s prestacion=%r",
                    existing["db_id"], canal_nuevo, fuente_nueva, prestacion_update,
                )
            except Exception as e_upd:
                log.warning("prellenar_pagos: error UPDATE id=%s: %s", existing["db_id"], e_upd)
                errores += 1
            continue

        # Fila nueva: INSERT draft
        id_prof   = cita.get("id_profesional")
        prof_info = PROFESIONALES.get(id_prof, {}) if id_prof else {}

        nombre_pac  = (cita.get("nombre_paciente") or "").strip()
        profesional = _nombre_profesional_pagos(
            id_prof, cita.get("nombre_profesional") or prof_info.get("nombre", "")
        )
        area        = prof_info.get("especialidad", "")
        hora_inicio = (cita.get("hora_inicio") or "")[:5]

        prestacion = ""
        if id_aten:
            prestacion = await _fetch_prestacion(id_aten)

        # Previsión + precio desde la atención de Medilink (convenio → fonasa/particular).
        convenio, _total = await _fetch_atencion_meta(id_aten)
        prevision_cita = _convenio_a_prevision(convenio)
        sug = _sugerir_copago(area, prevision_cita, id_prof)
        copago_sug = int(sug.get("copago_sugerido") or 0)

        # Derivar atribución (canal/fuente/confianza). El teléfono de la ficha de
        # Medilink hace el cruce EXACTO; rut_cita ya viene relleno desde la ficha.
        att = _resolver_atribucion(
            id_cita=id_cita_str, rut=rut_cita, nombre=nombre_pac,
            telefono_cita=tel_ficha,
        )
        canal_cita, fuente_cita = att["canal"], att["fuente"]
        origen_cita = _canal_a_origen(canal_cita)
        rut_final = att["rut"]

        try:
            with _conn() as conn:
                conn.execute(
                    """INSERT INTO pagos_cmc
                       (fecha, hora, paciente_nombre, rut, id_profesional, profesional,
                        area, prevision, copago, bonificacion, metodo_pago, folio,
                        codigo_transferencia, tipo_bono, procedimiento, origen, id_cita,
                        creado_por, bloqueado, canal, fuente, match_confianza, monto_medilink,
                        created_at, updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,0,'','','','',?,?,?,
                               'prellenar', 0, ?, ?, ?, ?, datetime('now'), datetime('now'))""",
                    (
                        fecha_iso,
                        hora_inicio,
                        nombre_pac,
                        rut_final,
                        id_prof,
                        profesional,
                        area,
                        prevision_cita,
                        copago_sug,
                        prestacion,
                        origen_cita,
                        id_cita_str,
                        canal_cita,
                        fuente_cita,
                        att["confianza"],
                        int(_total or 0),   # monto real Medilink (arancel) — Fase B cuadre
                    )
                )
                conn.commit()
            # Registrar en mapas para idempotencia dentro del mismo batch
            meta_new = {"db_id": None, "procedimiento": prestacion, "bloqueado": 0, "canal": canal_cita}
            if id_cita_str:
                existing_by_id_cita[id_cita_str] = meta_new
            if rut_cita:
                existing_by_rut[rut_cita] = meta_new
            creadas += 1
        except Exception as e_ins:
            log.warning("prellenar_pagos: error INSERT cita %s: %s", id_cita_str, e_ins)
            errores += 1

    log.info(
        "prellenar_pagos: fecha=%s creadas=%d actualizadas=%d saltadas=%d errores=%d",
        fecha_iso, creadas, actualizadas, saltadas, errores,
    )
    return {"creadas": creadas, "actualizadas": actualizadas, "saltadas": saltadas, "errores": errores}


@router.patch("/{pago_id}/lock")
async def toggle_lock_pago(
    pago_id: int,
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    """
    Alterna el campo bloqueado de una fila (0→1 o 1→0).
    Una fila bloqueada queda excluida de 'Actualizar desde Medilink'.
    """
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    from session import _conn
    with _conn() as conn:
        row = conn.execute(
            "SELECT id, COALESCE(bloqueado, 0) as bloqueado FROM pagos_cmc WHERE id = ?",
            (pago_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, f"Pago {pago_id} no encontrado")
        nuevo_estado = 0 if row["bloqueado"] else 1
        conn.execute(
            "UPDATE pagos_cmc SET bloqueado = ?, updated_at = datetime('now') WHERE id = ?",
            (nuevo_estado, pago_id)
        )
        conn.commit()

    log.info("pagos_routes.toggle_lock_pago: id=%d bloqueado=%d", pago_id, nuevo_estado)
    return {"ok": True, "id": pago_id, "bloqueado": nuevo_estado}


@router.delete("/{pago_id}")
async def delete_pago(
    pago_id: int,
    request: Request,
    token: str | None = Query(None),
    cmc_session: str | None = Cookie(None),
):
    """Elimina un registro de pago (errores de entrada / duplicados)."""
    _require_admin_dep(request, token=token, cmc_session=cmc_session)
    ensure_pagos_table()

    from session import _conn
    with _conn() as conn:
        exists = conn.execute(
            "SELECT id FROM pagos_cmc WHERE id = ?", (pago_id,)
        ).fetchone()
        if not exists:
            raise HTTPException(404, f"Pago {pago_id} no encontrado")
        conn.execute("DELETE FROM pagos_cmc WHERE id = ?", (pago_id,))
        conn.commit()

    log.info("pagos_routes.delete_pago: id=%d", pago_id)
    return {"ok": True, "id": pago_id}
