#!/usr/bin/env python3
"""
Importa la hoja 'Diario' del Google Sheet histórico 'pacientes diarios (1).xlsx'
a la tabla pagos_cmc de sessions.db (SQLCipher) en el VPS.

Uso local (desde Mac, genera SQL):
    python3 scripts/import_sheet_diario.py --dry-run

Uso en VPS (inserta directo):
    python3 /opt/chatbot-cmc/scripts/import_sheet_diario.py --execute
"""
import argparse
import os
import sys
import datetime
import re
import unicodedata

# ── Constantes ─────────────────────────────────────────────────────────────

XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    "/opt/chatbot-cmc/scripts/pacientes_diarios_import.xlsx"
)
DB_PATH   = "/opt/chatbot-cmc/data/sessions.db"
# Llave de cifrado SQLCipher: SOLO desde el entorno (igual que app/session.py).
# Nunca hardcodear — es la llave de la DB con datos clínicos (Ley 21.719).
# En el VPS: `set -a; source /opt/chatbot-cmc/.env; set +a` antes de ejecutar.
SQLCIPHER_KEY = (os.environ.get("SQLCIPHER_KEY") or "").strip()

# Dict PROFESIONALES copiado de medilink.py para resolución de id_profesional
PROFESIONALES = {
     1: "Dr. Rodrigo Olavarría",
    73: "Dr. Andrés Abarca",
    13: "Dr. Alonso Márquez",
    23: "Dr. Manuel Borrego",
    60: "Dr. Miguel Millán",
    61: "Dr. Tirso Rejón",
    65: "Dr. Nicolás Quijano",
    55: "Dra. Javiera Burgos",
    72: "Dr. Carlos Jiménez",
    66: "Dra. Daniela Castillo",
    75: "Dr. Fernando Fredes",
    69: "Dra. Aurora Valdés",
    76: "Dra. Valentina Fuentealba",
    59: "Paola Acosta",
    77: "Luis Armijo",
    21: "Leonardo Etcheverry",
    52: "Gisela Pinto",
    74: "Jorge Montalba",
    49: "Juan Pablo Rodríguez",
    70: "Juana Arratia",
    67: "Sarai Gómez",
    56: "Andrea Guevara",
    68: "David Pardo",
}

def _normalize(text: str) -> str:
    """Minúsculas, sin tildes, sin espacios extra."""
    if not text:
        return ""
    t = unicodedata.normalize("NFD", str(text).lower().strip())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t)

def _build_name_map():
    """Construye mapa de variaciones de nombre → id_profesional."""
    m = {}
    for pid, full_name in PROFESIONALES.items():
        # Nombre completo normalizado
        norm = _normalize(full_name)
        m[norm] = pid
        # Sin título (Dr./Dra.)
        without_title = re.sub(r'^(dr\.|dra\.)\s*', '', norm).strip()
        m[without_title] = pid
        # Solo apellido (primer token del nombre-sin-titulo)
        parts = without_title.split()
        if parts:
            m[parts[0]] = pid  # primer nombre/apellido
        if len(parts) > 1:
            m[parts[-1]] = pid  # último apellido
    return m

_NAME_MAP = _build_name_map()

def _resolve_profesional(raw: str):
    """Devuelve (id_profesional, nombre_normalizado). id puede ser None."""
    if not raw or not raw.strip():
        return None, ""
    norm = _normalize(raw)
    # Búsqueda exacta primero
    if norm in _NAME_MAP:
        pid = _NAME_MAP[norm]
        return pid, PROFESIONALES[pid]
    # Búsqueda parcial: ¿algún token del nombre está en el mapa?
    for token in norm.split():
        if len(token) > 3 and token in _NAME_MAP:
            pid = _NAME_MAP[token]
            return pid, PROFESIONALES[pid]
    # Sin match — devuelve el nombre original limpio
    return None, raw.strip()

def _normalize_prevision(raw) -> str:
    if not raw:
        return "particular"
    r = _normalize(str(raw))
    if "fonasa" in r:
        return "fonasa"
    return "particular"

def _normalize_metodo_pago(raw) -> str:
    if not raw:
        return "efectivo"
    r = _normalize(str(raw))
    # typos conocidos: "efectiivo", "debiito"
    r = r.replace("ii", "i")
    if "transfer" in r:
        return "transferencia"
    if "debito" in r or "débito" in r or "debit" in r:
        return "debito"
    if "credito" in r or "crédito" in r or "credit" in r:
        return "credito"
    return "efectivo"

def _parse_valor(raw) -> int:
    if raw is None:
        return 0
    try:
        return int(float(str(raw).replace(",", "").replace("$", "").strip()))
    except Exception:
        return 0

def _parse_hora(raw) -> str:
    if raw is None:
        return "00:00"
    if isinstance(raw, datetime.time):
        return raw.strftime("%H:%M")
    s = str(raw).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    return s[:5]

def _normalize_area(raw) -> str:
    if not raw:
        return ""
    return str(raw).strip().rstrip()


def parse_xlsx(xlsx_path: str):
    """Parsea hoja 'Diario' y retorna lista de dicts listos para insertar."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)

    results = []
    unmapped = []

    # Solo hoja "Diario"
    ws = wb["Diario"]

    # Fecha del día: celda B1
    fecha_raw = ws["B1"].value
    if isinstance(fecha_raw, str):
        # formato DD/MM/YYYY
        parts = fecha_raw.strip().split("/")
        fecha_iso = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    elif isinstance(fecha_raw, datetime.datetime):
        fecha_iso = fecha_raw.strftime("%Y-%m-%d")
    else:
        raise ValueError(f"No se pudo parsear la fecha del sheet: {fecha_raw!r}")

    print(f"Fecha del sheet: {fecha_iso}")

    # Headers en fila 3: HORA, PACIENTE, NOMBRE PROFESIONAL, AREA, PREVISION, VALOR, METODO DE PAGO, N° FOLIO, PROCEDIMIENTO
    # Datos desde fila 4
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Skip filas vacías
        if not any(v is not None and str(v).strip() not in ("", "None") for v in row[:9]):
            continue

        hora     = _parse_hora(row[0])
        paciente = str(row[1] or "").strip()
        prof_raw = str(row[2] or "").strip()
        area     = _normalize_area(row[3])
        prevision = _normalize_prevision(row[4])
        valor    = _parse_valor(row[5])
        metodo   = _normalize_metodo_pago(row[6])
        folio    = str(int(float(row[7]))) if row[7] is not None and str(row[7]).strip() not in ("", "None") else ""
        proc     = str(row[8] or "").strip()

        if not paciente:
            continue

        id_prof, prof_nombre = _resolve_profesional(prof_raw)

        if id_prof is None and prof_raw:
            unmapped.append(f"  fila {row_idx}: '{prof_raw}' → sin ID (guardado como '{prof_nombre}')")

        results.append({
            "fecha":           fecha_iso,
            "hora":            hora,
            "paciente_nombre": paciente,
            "rut":             "",
            "id_profesional":  id_prof,
            "profesional":     prof_nombre,
            "area":            area,
            "prevision":       prevision,
            "copago":          valor,
            "bonificacion":    0,
            "metodo_pago":     metodo,
            "folio":           folio,
            "codigo_transferencia": "",
            "procedimiento":   proc,
            "origen":          "import_sheet",
            "id_cita":         "",
            "creado_por":      "import_sheet",
        })

    return fecha_iso, results, unmapped


def insert_rows(rows: list, dry_run: bool = False):
    """Inserta en pagos_cmc usando INSERT OR IGNORE para idempotencia."""
    if dry_run:
        print(f"\n[DRY-RUN] Se insertarían {len(rows)} filas:")
        for r in rows:
            print(f"  {r['hora']} | {r['paciente_nombre'][:30]:<30} | {r['profesional'][:25]:<25} | {r['prevision']:<10} | {r['copago']:>7} | {r['metodo_pago']:<15} | folio={r['folio']}")
        return

    # Conectar con SQLCipher — cifrado OBLIGATORIO (fail-closed).
    # Nunca caer a sqlite3 plano: escribiría PII clínica sin cifrar.
    if not SQLCIPHER_KEY:
        sys.exit("ERROR: SQLCIPHER_KEY no está en el entorno. Corre "
                 "`set -a; source /opt/chatbot-cmc/.env; set +a` antes de ejecutar con --execute.")
    if not re.fullmatch(r"[0-9a-fA-F]+", SQLCIPHER_KEY):
        sys.exit("ERROR: SQLCIPHER_KEY debe ser hexadecimal (0-9, a-f).")
    try:
        from sqlcipher3 import dbapi2 as sc
    except ImportError:
        sys.exit("ERROR: sqlcipher3 no disponible — abortando para no escribir PII en claro. "
                 "Instala sqlcipher3 en el VPS, o usa --dry-run para validar sin tocar la DB.")
    conn = sc.connect(DB_PATH, timeout=10)
    conn.execute(f"PRAGMA key = \"x'{SQLCIPHER_KEY}'\"")

    conn.execute("PRAGMA journal_mode=WAL")

    # Idempotencia: unicidad por fecha+hora+paciente_nombre+folio
    # Si folio está vacío, usamos fecha+hora+paciente_nombre+profesional
    SQL = """
        INSERT OR IGNORE INTO pagos_cmc
            (fecha, hora, paciente_nombre, rut, id_profesional, profesional,
             area, prevision, copago, bonificacion, metodo_pago, folio,
             codigo_transferencia, procedimiento, origen, id_cita,
             creado_por, created_at, updated_at)
        SELECT ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'),datetime('now')
        WHERE NOT EXISTS (
            SELECT 1 FROM pagos_cmc
            WHERE fecha = ?
              AND hora  = ?
              AND paciente_nombre = ?
              AND (folio = ? OR (folio = '' AND profesional = ?))
        )
    """

    inserted = 0
    skipped  = 0
    for r in rows:
        cur = conn.execute(SQL, (
            r["fecha"], r["hora"], r["paciente_nombre"], r["rut"],
            r["id_profesional"], r["profesional"], r["area"], r["prevision"],
            r["copago"], r["bonificacion"], r["metodo_pago"], r["folio"],
            r["codigo_transferencia"], r["procedimiento"], r["origen"],
            r["id_cita"], r["creado_por"],
            # WHERE params:
            r["fecha"], r["hora"], r["paciente_nombre"], r["folio"], r["profesional"],
        ))
        if cur.rowcount > 0:
            inserted += 1
        else:
            skipped += 1

    conn.commit()
    conn.close()
    print(f"Insertadas: {inserted} | Ya existían (skip): {skipped}")
    return inserted, skipped


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Solo mostrar, no insertar")
    parser.add_argument("--execute", action="store_true", help="Insertar en DB")
    parser.add_argument("--xlsx", default=XLSX_PATH, help="Ruta al xlsx")
    args = parser.parse_args()

    if not args.dry_run and not args.execute:
        print("Usar --dry-run o --execute")
        sys.exit(1)

    xlsx = args.xlsx
    if not os.path.exists(xlsx):
        print(f"ERROR: No se encuentra {xlsx}")
        sys.exit(1)

    fecha_iso, rows, unmapped = parse_xlsx(xlsx)
    print(f"Filas parseadas: {len(rows)}")

    if unmapped:
        print(f"\nProfesionales sin ID Medilink ({len(unmapped)}):")
        for u in unmapped:
            print(u)

    insert_rows(rows, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
